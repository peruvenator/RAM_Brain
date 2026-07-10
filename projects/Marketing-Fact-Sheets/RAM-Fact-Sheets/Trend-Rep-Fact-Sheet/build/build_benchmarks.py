"""Build trend-rep-charts.xlsx — equity curves (base 100) plus full proof-of-calculation sheets.

Series (equity curves, base 100 at first live date):
  - Trend Replication Live     (from data/trend-rep-live-performance.xlsx)
  - SocGen Trend Index (Excess) (NEIXCTAT_US_Index monthly return minus LD12TRUU monthly return)
  - Bottom Up (Net)            (sub_strategies/bottom_up/returns.csv net_fees compounded monthly)
  - Top Down (Med, Net)        (sub_strategies/top_down_med/returns.csv)
  - Top Down (Small, Net)      (sub_strategies/top_down_small/returns.csv)

Update strategy: loads the existing xlsx if present (preserves any embedded chart),
clears and rewrites the data-bearing sheets, then saves. Creates the file from scratch
if it doesn't exist.

Sheets written:
  - Equity (Base 100)        : 5 series rebased to 100 at first live date
  - Monthly Returns          : same 5 series as monthly returns
  - SocGen Calc              : month-end closes + monthly returns for NEIXCTAT and LD12TRUU,
                               with excess return = NEIXCTAT return - LD12TRUU return
                               (Excel formulas so the math is auditable in-place)
  - Daily - Bottom Up        : daily net_fees returns, Month tag, within-month compounding
  - Daily - Top Down Med     : same
  - Daily - Top Down Small   : same
"""
import argparse
import re
import shutil
import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
from pandas.tseries.holiday import USFederalHolidayCalendar
from pandas.tseries.offsets import CustomBusinessDay
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
LIVE_PATH = PROJECT_ROOT / "data" / "trend-rep-live-performance.xlsx"
OUT_PATH  = PROJECT_ROOT / "data" / "trend-rep-charts.xlsx"
ARCHIVE_DIR = PROJECT_ROOT / "data" / "archive"

S3 = Path(r"C:/Users/RodrigoGordillo/S3_Data")
SOCGEN_PATH = S3 / "trading_data/etfs/ratio_adjusted/NEIXCTAT_US_Index.csv"
CASH_PATH   = S3 / "trading_data/etfs/ratio_adjusted/LD12TRUU_Index.csv"
SUB_STRATS = {
    "Bottom Up (Net)":       S3 / "backtest-sims/portfolio_returns/trend_rep/rebuild/sub_strategies/bottom_up/returns.csv",
    "Top Down (Med, Net)":   S3 / "backtest-sims/portfolio_returns/trend_rep/rebuild/sub_strategies/top_down_med/returns.csv",
    "Top Down (Small, Net)": S3 / "backtest-sims/portfolio_returns/trend_rep/rebuild/sub_strategies/top_down_small/returns.csv",
}

SERIES_COLUMNS = [
    "Trend Replication Live",
    "SocGen Trend Index (Excess)",
    "Bottom Up (Net)",
    "Top Down (Med, Net)",
    "Top Down (Small, Net)",
]

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
NOTE_FONT   = Font(italic=True, color="555555")


# ---------- loaders ----------

def load_live():
    df = pd.read_excel(LIVE_PATH)
    df.columns = ["Date", "Trend Replication Live"]
    df["Date"] = pd.to_datetime(df["Date"])
    df["Month"] = df["Date"].dt.to_period("M")
    return df


def monthly_close(path: Path) -> pd.Series:
    df = pd.read_csv(path, parse_dates=["date"])
    df["Month"] = df["date"].dt.to_period("M")
    return df.groupby("Month")["close"].last()


def sub_strategy_daily(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, parse_dates=["date"])
    df["Month"] = df["date"].dt.to_period("M")
    return df[["date", "Month", "net_fees"]].rename(columns={"date": "Date", "net_fees": "Daily Return (net_fees)"})


def sub_strategy_monthly_returns(daily: pd.DataFrame) -> pd.Series:
    return daily.groupby("Month")["Daily Return (net_fees)"].apply(lambda x: (1 + x).prod() - 1)


# ---------- equity curve ----------

def to_equity(returns: pd.Series, base: float = 100.0) -> list:
    vals = [base]
    for r in returns.iloc[1:]:
        r = 0.0 if pd.isna(r) else r
        vals.append(vals[-1] * (1 + r))
    return vals


# ---------- sheet helpers ----------

def ensure_sheet(wb, name):
    """Clear an existing sheet's values (preserving charts anchored to it) or create new."""
    if name in wb.sheetnames:
        ws = wb[name]
        max_row = ws.max_row
        max_col = ws.max_column
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.value = None
                cell.number_format = "General"
    else:
        ws = wb.create_sheet(name)
    return ws


def replace_sheet(wb, name):
    """Remove and recreate a sheet (for calc sheets, which the user's chart does not reference)."""
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def style_header_row(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- sheet writers ----------

def update_chart_ranges(ws, n_data_rows):
    """Rewrite each chart's series ranges on this sheet so they span the new data extent.

    The chart's formatting (colors, line styles, title, legend, axis settings) is untouched —
    only the cell references for data, categories, and series titles are bumped.
    """
    end_row = n_data_rows + 1
    row_pattern = re.compile(r"\$\d+$")

    def bump(ref: str) -> str:
        if ":" not in ref:
            return ref
        left, right = ref.rsplit(":", 1)
        return f"{left}:{row_pattern.sub(f'${end_row}', right)}"

    for chart in getattr(ws, "_charts", []):
        for ser in getattr(chart, "series", []):
            if ser.val and ser.val.numRef and ser.val.numRef.f:
                ser.val.numRef.f = bump(ser.val.numRef.f)
            if ser.cat:
                if ser.cat.numRef and ser.cat.numRef.f:
                    ser.cat.numRef.f = bump(ser.cat.numRef.f)
                elif ser.cat.strRef and ser.cat.strRef.f:
                    ser.cat.strRef.f = bump(ser.cat.strRef.f)


def write_equity_sheet(ws, equity_df):
    headers = list(equity_df.columns)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, len(headers))
    for r, row in enumerate(equity_df.itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.number_format = "yyyy-mm-dd"
            else:
                cell.number_format = "0.00"
    set_col_widths(ws, [12] + [22] * (len(headers) - 1))
    ws.freeze_panes = "B2"


def write_monthly_returns_sheet(ws, returns_df):
    headers = list(returns_df.columns)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, len(headers))
    for r, row in enumerate(returns_df.itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.number_format = "yyyy-mm-dd"
            else:
                cell.number_format = "0.00%"
    set_col_widths(ws, [12] + [22] * (len(headers) - 1))
    ws.freeze_panes = "B2"


def write_socgen_calc_sheet(ws, socgen_close: pd.Series, cash_close: pd.Series, live_months):
    """Month-end closes for NEIXCTAT and LD12TRUU with Excel formulas for monthly + excess returns.

    Layout:
      A: Month (display as yyyy-mm)
      B: NEIXCTAT close (month-end total-return index level)
      C: NEIXCTAT monthly return (formula = B_n/B_{n-1} - 1)
      D: LD12TRUU close (month-end cash index level)
      E: LD12TRUU monthly return (formula = D_n/D_{n-1} - 1)
      F: SocGen Excess = C - E (formula)

    Includes one extra prior-month row (the month before the first live-file month) so the
    first live-file month has a valid return denominator.
    """
    live_months_list = list(live_months)
    first_month = live_months_list[0]
    prior_month = first_month - 1

    months = [prior_month] + live_months_list

    headers = [
        "Month",
        "NEIXCTAT Close (Total Return)",
        "NEIXCTAT Monthly Return",
        "LD12TRUU Close (Cash TR)",
        "LD12TRUU Monthly Return",
        "SocGen Trend Index (Excess) = NEIXCTAT - LD12TRUU",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, len(headers))

    for i, m in enumerate(months):
        row = i + 2
        ws.cell(row=row, column=1, value=str(m))
        ws.cell(row=row, column=2, value=float(socgen_close.get(m, float("nan"))))
        ws.cell(row=row, column=4, value=float(cash_close.get(m, float("nan"))))
        if i > 0:
            ws.cell(row=row, column=3, value=f"=B{row}/B{row-1}-1")
            ws.cell(row=row, column=5, value=f"=D{row}/D{row-1}-1")
            ws.cell(row=row, column=6, value=f"=C{row}-E{row}")

    for r in range(2, len(months) + 2):
        ws.cell(row=r, column=2).number_format = "0.0000"
        ws.cell(row=r, column=4).number_format = "0.0000"
        ws.cell(row=r, column=3).number_format = "0.00%"
        ws.cell(row=r, column=5).number_format = "0.00%"
        ws.cell(row=r, column=6).number_format = "0.00%"

    set_col_widths(ws, [12, 30, 22, 26, 22, 48])
    ws.freeze_panes = "B2"

    note_row = len(months) + 4
    ws.cell(row=note_row, column=1,
            value="Note: Monthly returns are computed from month-end index closes (last trading day of each month). "
                  "Excess return = NEIXCTAT monthly return - LD12TRUU monthly return. "
                  "The first data row (prior month) exists only to serve as the denominator for the first return.")
    ws.cell(row=note_row, column=1).font = NOTE_FONT
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)


def write_daily_sheet(ws, daily: pd.DataFrame, label: str, live_months):
    """Daily returns for a sub-strategy with in-month running wealth and month-end monthly return.

    Layout:
      A: Date
      B: Daily Return (net_fees)
      C: Month
      D: Running Wealth in Month (formula: resets each month start, compounds each day)
      E: Monthly Return (populated on last observation of each month, formula = D - 1)
    """
    live_set = set(live_months)
    first_month = min(live_months)
    prior_month = first_month - 1
    filtered = daily[daily["Month"].isin([prior_month] + list(live_set))].copy()
    filtered = filtered.sort_values("Date").reset_index(drop=True)

    headers = [
        "Date",
        f"Daily Return (net_fees) - {label}",
        "Month",
        "Running Wealth in Month",
        "Monthly Return (month-end only)",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, len(headers))

    n = len(filtered)
    months_col = filtered["Month"].astype(str).tolist()
    for i, (_, r) in enumerate(filtered.iterrows()):
        row = i + 2
        ws.cell(row=row, column=1, value=r["Date"].to_pydatetime() if hasattr(r["Date"], "to_pydatetime") else r["Date"])
        ws.cell(row=row, column=2, value=float(r["Daily Return (net_fees)"]))
        ws.cell(row=row, column=3, value=str(r["Month"]))

        if i == 0:
            ws.cell(row=row, column=4, value=f"=1+B{row}")
        else:
            ws.cell(row=row, column=4, value=f'=IF(C{row}=C{row-1}, D{row-1}*(1+B{row}), 1+B{row})')

        is_month_end = (i == n - 1) or (months_col[i] != months_col[i + 1])
        if is_month_end:
            ws.cell(row=row, column=5, value=f"=D{row}-1")

    for row in range(2, n + 2):
        ws.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=2).number_format = "0.0000%"
        ws.cell(row=row, column=4).number_format = "0.000000"
        ws.cell(row=row, column=5).number_format = "0.00%"

    set_col_widths(ws, [12, 34, 10, 24, 28])
    ws.freeze_panes = "A2"

    note_row = n + 3
    ws.cell(row=note_row, column=1,
            value=f"Note: Monthly return for {label} = product of (1 + daily net_fees) across all trading days "
                  f"in the calendar month, minus 1. Column D shows the running in-month wealth (reset on first trading "
                  f"day of each month). Column E shows the final monthly return on the last trading day of each month.")
    ws.cell(row=note_row, column=1).font = NOTE_FONT
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)


# ---------- gates ----------

def is_nth_business_day(n: int, when: datetime | None = None) -> bool:
    """Return True if `when` is the Nth US business day of its month.

    Business days = weekdays excluding US federal holidays (pandas USFederalHolidayCalendar).
    """
    when = when or datetime.now()
    today = pd.Timestamp(when.date())
    month_start = today.replace(day=1)
    cbd = CustomBusinessDay(calendar=USFederalHolidayCalendar())
    bdays = pd.date_range(start=month_start, end=today, freq=cbd)
    return len(bdays) == n and bdays[-1].date() == today.date()


def live_file_has_prior_month(live_df: pd.DataFrame, when: datetime | None = None) -> bool:
    """Return True if the live file contains at least one row in the prior calendar month."""
    when = when or datetime.now()
    prior_month = (pd.Timestamp(when.date()) - pd.offsets.MonthBegin(1)).to_period("M")
    return prior_month in set(live_df["Month"].tolist())


def archive_existing_output():
    """Copy current xlsx to data/archive/trend-rep-charts_{YYYYMM}.xlsx before rebuild."""
    if not OUT_PATH.exists():
        return None
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m")
    dest = ARCHIVE_DIR / f"trend-rep-charts_{stamp}.xlsx"
    shutil.copy2(OUT_PATH, dest)
    return dest


# ---------- orchestration ----------

def main(gate_business_day: int | None = None, force: bool = False):
    if gate_business_day is not None and not force:
        if not is_nth_business_day(gate_business_day):
            print(f"[skip] today ({datetime.now().strftime('%Y-%m-%d %a')}) "
                  f"is not business day #{gate_business_day} of the month. Exiting cleanly.")
            return 0

    live = load_live()
    live_months = live["Month"].tolist()

    if gate_business_day is not None and not force:
        if not live_file_has_prior_month(live):
            print(f"[error] live file {LIVE_PATH.name} has no data for the prior calendar month. "
                  f"Upstream update has not landed yet. Aborting.", file=sys.stderr)
            return 1

    archived = archive_existing_output()
    if archived:
        print(f"[info] archived previous output to {archived}")

    socgen_close = monthly_close(SOCGEN_PATH)
    cash_close   = monthly_close(CASH_PATH)

    socgen_mret = socgen_close.pct_change()
    cash_mret   = cash_close.pct_change()
    socgen_excess_mret = socgen_mret - cash_mret

    bu_daily  = sub_strategy_daily(SUB_STRATS["Bottom Up (Net)"])
    tdm_daily = sub_strategy_daily(SUB_STRATS["Top Down (Med, Net)"])
    tds_daily = sub_strategy_daily(SUB_STRATS["Top Down (Small, Net)"])

    bu_mret  = sub_strategy_monthly_returns(bu_daily)
    tdm_mret = sub_strategy_monthly_returns(tdm_daily)
    tds_mret = sub_strategy_monthly_returns(tds_daily)

    returns_df = pd.DataFrame({"Date": live["Date"].values})
    returns_df["Trend Replication Live"]       = live["Trend Replication Live"].values
    returns_df["SocGen Trend Index (Excess)"]  = live["Month"].map(socgen_excess_mret).values
    returns_df["Bottom Up (Net)"]              = live["Month"].map(bu_mret).values
    returns_df["Top Down (Med, Net)"]          = live["Month"].map(tdm_mret).values
    returns_df["Top Down (Small, Net)"]        = live["Month"].map(tds_mret).values

    missing = returns_df[SERIES_COLUMNS].isna().sum()
    if missing.any():
        print("[warn] missing monthly returns per series:")
        print(missing[missing > 0])

    equity_df = returns_df[["Date"]].copy()
    for col in SERIES_COLUMNS:
        equity_df[col] = to_equity(returns_df[col])

    # --- open or create workbook ---
    if OUT_PATH.exists():
        print(f"[info] loading existing workbook: {OUT_PATH}")
        wb = load_workbook(OUT_PATH)
    else:
        print(f"[info] creating new workbook: {OUT_PATH}")
        wb = Workbook()
        default = wb.active
        wb.remove(default)

    # Equity and Monthly Returns sheets: update in place so any chart anchored here survives.
    eq_ws  = ensure_sheet(wb, "Equity (Base 100)")
    write_equity_sheet(eq_ws, equity_df)
    update_chart_ranges(eq_ws, n_data_rows=len(equity_df))

    mr_ws  = ensure_sheet(wb, "Monthly Returns")
    write_monthly_returns_sheet(mr_ws, returns_df)

    # Calc sheets: recreated each run (user chart is not on these).
    sg_ws  = replace_sheet(wb, "SocGen Calc")
    write_socgen_calc_sheet(sg_ws, socgen_close, cash_close, live_months)

    write_daily_sheet(replace_sheet(wb, "Daily - Bottom Up"),       bu_daily,  "Bottom Up (Net)",       live_months)
    write_daily_sheet(replace_sheet(wb, "Daily - Top Down Med"),    tdm_daily, "Top Down (Med, Net)",   live_months)
    write_daily_sheet(replace_sheet(wb, "Daily - Top Down Small"),  tds_daily, "Top Down (Small, Net)", live_months)

    desired_order = [
        "Equity (Base 100)",
        "Monthly Returns",
        "SocGen Calc",
        "Daily - Bottom Up",
        "Daily - Top Down Med",
        "Daily - Top Down Small",
    ]
    present = [s for s in desired_order if s in wb.sheetnames]
    other = [s for s in wb.sheetnames if s not in desired_order]
    wb._sheets = [wb[s] for s in present + other]

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"[ok] wrote {OUT_PATH} at {datetime.now().isoformat(timespec='seconds')}")
    print(f"[info] equity curve last month: {equity_df.iloc[-1].to_dict()}")
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Rebuild trend-rep-charts.xlsx")
    parser.add_argument("--gate-business-day", type=int, default=None,
                        help="Only run if today is the Nth US business day of the month (skip cleanly otherwise).")
    parser.add_argument("--force", action="store_true",
                        help="Bypass business-day and live-file gates.")
    args = parser.parse_args()
    try:
        sys.exit(main(gate_business_day=args.gate_business_day, force=args.force))
    except Exception as exc:
        print(f"[error] unhandled exception: {exc!r}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
