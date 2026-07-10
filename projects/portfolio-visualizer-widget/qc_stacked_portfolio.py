"""
QC Spreadsheet: 100% Global Equities + 100% Managed Futures (Stacked)
Replicates the widget's computePortfolio() logic with all formulas exposed.

Portfolio:
  Core:  100% Global Equities (MSCI ACWI)
  Stack: 100% Managed Futures (MFT)
  Fees:  All zero (advisory=0, feeBp=0, financingBp=0)
  Financing: T-Bill base deduction still applies per widget logic
"""

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── Load data ──
with open("data/indices_compact.json", "r") as f:
    data = json.load(f)

dates = data["dates"]
acwi = data["series"]["MSCI ACWI"]
mft = data["series"]["MFT"]
tb = data["series"]["Treasury Bill"]

# ── Common date range ──
common_start = max(acwi["start"], mft["start"], tb["start"])
common_end = min(
    acwi["start"] + len(acwi["values"]) - 1,
    mft["start"] + len(mft["values"]) - 1,
    tb["start"] + len(tb["values"]) - 1,
)
num_months = common_end - common_start  # number of return periods

# ── Create workbook ──
wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(name="DM Sans", bold=True, size=11)
header_fill = PatternFill(start_color="323A46", end_color="323A46", fill_type="solid")
header_font_white = Font(name="DM Sans", bold=True, size=11, color="FFFFFF")
title_font = Font(name="DM Sans", bold=True, size=14)
pct_fmt = '0.00%'
pct_fmt_4 = '0.0000%'
num_fmt = '#,##0.0000'
idx_fmt = '#,##0.00'
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

def style_data_cell(cell, fmt=None):
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")
    if fmt:
        cell.number_format = fmt

# ════════════════════════════════════════
# SHEET 1: Monthly Data
# ════════════════════════════════════════
ws = wb.active
ws.title = "Monthly Data"

# Title
ws.cell(row=1, column=1, value="QC: 100% Global Equities + 100% Managed Futures (Stacked)").font = title_font
ws.cell(row=2, column=1, value="Core: 100% MSCI ACWI  |  Stack: 100% MFT (SG CTA)  |  All fees = 0  |  T-Bill financing deducted from overlay").font = Font(name="DM Sans", size=10, italic=True)

# Headers (row 4)
headers = [
    "Date",                           # A
    "MSCI ACWI\n(Index Level)",       # B
    "MFT\n(Index Level)",             # C
    "Treasury Bill\n(Index Level)",   # D
    "Global Equities\nReturn",        # E
    "Managed Futures\nReturn",        # F
    "T-Bill\nReturn",                 # G
    "Core Return\n(100% ACWI)",       # H
    "Overlay Net Return\n(MF - T-Bill)", # I
    "Stacked Return\n(Core + Overlay)", # J
    "Growth of $1\n(Core)",           # K
    "Growth of $1\n(Stacked)",        # L
]
HEADER_ROW = 4
for c, h in enumerate(headers, 1):
    ws.cell(row=HEADER_ROW, column=c, value=h)
style_header(ws, HEADER_ROW, len(headers))

# Column widths
col_widths = [12, 16, 14, 16, 16, 16, 14, 16, 20, 20, 16, 16]
for c, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

# ── Write data with FORMULAS ──
# Row 5 = first date (index level only, no return yet)
# Row 6 onward = returns computed from index levels

data_start_row = 5

# First row: just the starting index levels and Growth = 1
r = data_start_row
dt = dates[common_start]
acwi_idx = acwi["values"][common_start - acwi["start"]]
mft_idx = mft["values"][common_start - mft["start"]]
tb_idx = tb["values"][common_start - tb["start"]]

ws.cell(row=r, column=1, value=dt)
style_data_cell(ws.cell(row=r, column=1))
ws.cell(row=r, column=2, value=acwi_idx)
style_data_cell(ws.cell(row=r, column=2), idx_fmt)
ws.cell(row=r, column=3, value=mft_idx)
style_data_cell(ws.cell(row=r, column=3), idx_fmt)
ws.cell(row=r, column=4, value=tb_idx)
style_data_cell(ws.cell(row=r, column=4), idx_fmt)

# No returns for first row
for c in range(5, 11):
    style_data_cell(ws.cell(row=r, column=c))

# Growth of $1 starts at 1
ws.cell(row=r, column=11, value=1)
style_data_cell(ws.cell(row=r, column=11), num_fmt)
ws.cell(row=r, column=12, value=1)
style_data_cell(ws.cell(row=r, column=12), num_fmt)

# Monthly return rows
for m in range(1, num_months + 1):
    r = data_start_row + m
    idx = common_start + m

    dt = dates[idx]
    a_val = acwi["values"][idx - acwi["start"]]
    m_val = mft["values"][idx - mft["start"]]
    t_val = tb["values"][idx - tb["start"]]

    # A: Date
    ws.cell(row=r, column=1, value=dt)
    style_data_cell(ws.cell(row=r, column=1))

    # B: ACWI index
    ws.cell(row=r, column=2, value=a_val)
    style_data_cell(ws.cell(row=r, column=2), idx_fmt)

    # C: MFT index
    ws.cell(row=r, column=3, value=m_val)
    style_data_cell(ws.cell(row=r, column=3), idx_fmt)

    # D: T-Bill index
    ws.cell(row=r, column=4, value=t_val)
    style_data_cell(ws.cell(row=r, column=4), idx_fmt)

    # E: Global Equities Return = B(r)/B(r-1) - 1
    ws.cell(row=r, column=5).value = f"=B{r}/B{r-1}-1"
    style_data_cell(ws.cell(row=r, column=5), pct_fmt_4)

    # F: Managed Futures Return = C(r)/C(r-1) - 1
    ws.cell(row=r, column=6).value = f"=C{r}/C{r-1}-1"
    style_data_cell(ws.cell(row=r, column=6), pct_fmt_4)

    # G: T-Bill Return = D(r)/D(r-1) - 1
    ws.cell(row=r, column=7).value = f"=D{r}/D{r-1}-1"
    style_data_cell(ws.cell(row=r, column=7), pct_fmt_4)

    # H: Core Return = 100% * E (same as E since 100% weight)
    ws.cell(row=r, column=8).value = f"=E{r}"
    style_data_cell(ws.cell(row=r, column=8), pct_fmt_4)

    # I: Overlay Net Return = MF Return - T-Bill Return
    # Widget formula: w * r[i] - w * cashReturns[i]  (w=1.0 here)
    ws.cell(row=r, column=9).value = f"=F{r}-G{r}"
    style_data_cell(ws.cell(row=r, column=9), pct_fmt_4)

    # J: Stacked Return = Core + Overlay Net
    ws.cell(row=r, column=10).value = f"=H{r}+I{r}"
    style_data_cell(ws.cell(row=r, column=10), pct_fmt_4)

    # K: Growth of $1 (Core) = K(r-1) * (1 + H(r))
    ws.cell(row=r, column=11).value = f"=K{r-1}*(1+H{r})"
    style_data_cell(ws.cell(row=r, column=11), num_fmt)

    # L: Growth of $1 (Stacked) = L(r-1) * (1 + J(r))
    ws.cell(row=r, column=12).value = f"=L{r-1}*(1+J{r})"
    style_data_cell(ws.cell(row=r, column=12), num_fmt)

last_data_row = data_start_row + num_months

# ════════════════════════════════════════
# Summary Statistics (below the data)
# ════════════════════════════════════════
stats_start = last_data_row + 3
ws.cell(row=stats_start, column=1, value="Summary Statistics").font = Font(name="DM Sans", bold=True, size=13)
ws.cell(row=stats_start + 1, column=1, value=f"Period: {dates[common_start]} to {dates[common_end]}  ({num_months} months)").font = Font(name="DM Sans", size=10, italic=True)

# Column labels for stats
sr = stats_start + 3
ws.cell(row=sr, column=1, value="Metric")
ws.cell(row=sr, column=2, value="Core\n(100% ACWI)")
ws.cell(row=sr, column=3, value="Stacked\n(ACWI + MF)")
ws.cell(row=sr, column=4, value="Formula Reference")
style_header(ws, sr, 4)
ws.column_dimensions["D"].width = 50

# Helper references
first_ret = data_start_row + 1  # first row with returns
last_ret = last_data_row
n_cells = f"{last_ret - first_ret + 1}"  # count of return months

# We'll use named ranges for clarity in formulas
# H = core returns, J = stacked returns, K = core growth, L = stacked growth
core_ret_range = f"H{first_ret}:H{last_ret}"
stack_ret_range = f"J{first_ret}:J{last_ret}"
tbill_ret_range = f"G{first_ret}:G{last_ret}"

stats = [
    ("Number of Months", f"=COUNT({core_ret_range})", f"=COUNT({stack_ret_range})", "COUNT of return cells"),
    ("Cumulative Return", f"=K{last_ret}/K{data_start_row}-1", f"=L{last_ret}/L{data_start_row}-1", "End Growth / Start Growth - 1"),
    ("Annualized Return", f"=(1+B{sr+2})^(12/B{sr+1})-1", f"=(1+C{sr+2})^(12/C{sr+1})-1", "(1 + CumRet)^(12/N) - 1"),
    ("Annualized Volatility", f"=STDEV({core_ret_range})*SQRT(12)", f"=STDEV({stack_ret_range})*SQRT(12)", "STDEV(monthly returns) * SQRT(12)"),
    ("Risk-Free Rate (Ann.)", f"=AVERAGE({tbill_ret_range})*12", f"=AVERAGE({tbill_ret_range})*12", "Average monthly T-Bill return * 12"),
    ("Sharpe Ratio", f"=(B{sr+3}-B{sr+5})/B{sr+4}", f"=(C{sr+3}-C{sr+5})/C{sr+4}", "(Ann Return - Rf) / Vol"),
    ("Downside Deviation (Ann.)",
     f"=SQRT(SUMPRODUCT((({core_ret_range}<0)*{core_ret_range})^2)/COUNT({core_ret_range}))*SQRT(12)",
     f"=SQRT(SUMPRODUCT((({stack_ret_range}<0)*{stack_ret_range})^2)/COUNT({stack_ret_range}))*SQRT(12)",
     "SQRT(sum of squared negative returns / N) * SQRT(12)"),
    ("Sortino Ratio", f"=B{sr+3}/B{sr+7}", f"=C{sr+3}/C{sr+7}", "Ann Return / Downside Dev"),
    ("Best Month", f"=MAX({core_ret_range})", f"=MAX({stack_ret_range})", "MAX of monthly returns"),
    ("Worst Month", f"=MIN({core_ret_range})", f"=MIN({stack_ret_range})", "MIN of monthly returns"),
    ("Average Monthly Return", f"=AVERAGE({core_ret_range})", f"=AVERAGE({stack_ret_range})", "AVERAGE of monthly returns"),
    ("Monthly Std Dev", f"=STDEV({core_ret_range})", f"=STDEV({stack_ret_range})", "STDEV of monthly returns"),
    ("Skewness", f"=SKEW({core_ret_range})", f"=SKEW({stack_ret_range})", "Excel SKEW (sample-adjusted)"),
    ("Excess Kurtosis", f"=KURT({core_ret_range})", f"=KURT({stack_ret_range})", "Excel KURT (excess, sample-adjusted)"),
]

for i, (label, core_formula, stack_formula, ref) in enumerate(stats):
    row = sr + 1 + i
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=1).font = Font(name="DM Sans", bold=True)
    ws.cell(row=row, column=1).border = thin_border

    ws.cell(row=row, column=2).value = core_formula
    ws.cell(row=row, column=3).value = stack_formula
    ws.cell(row=row, column=4, value=ref)
    ws.cell(row=row, column=4).font = Font(name="DM Sans", size=9, italic=True, color="666666")

    # Format
    if label == "Number of Months":
        fmt = '0'
    elif "Ratio" in label or "Skewness" in label or "Kurtosis" in label:
        fmt = '0.0000'
    else:
        fmt = pct_fmt

    for c in [2, 3]:
        style_data_cell(ws.cell(row=row, column=c), fmt)
    ws.cell(row=row, column=4).border = thin_border

# Note about Max Drawdown
dd_note_row = sr + 1 + len(stats) + 1
ws.cell(row=dd_note_row, column=1, value="Note: Max Drawdown requires a running peak calculation. See the 'Drawdown' sheet for the full formula chain.").font = Font(name="DM Sans", size=9, italic=True)

# ════════════════════════════════════════
# SHEET 2: Drawdown Calculation
# ════════════════════════════════════════
ws2 = wb.create_sheet("Drawdown")
ws2.cell(row=1, column=1, value="Max Drawdown Calculation").font = title_font

headers2 = ["Date", "Core Growth", "Core Peak", "Core Drawdown", "Stacked Growth", "Stacked Peak", "Stacked Drawdown"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=c, value=h)
style_header(ws2, 3, len(headers2))

for c in range(1, 8):
    ws2.column_dimensions[get_column_letter(c)].width = 16

# Reference Monthly Data sheet for growth values
for m in range(num_months + 1):
    r = 4 + m
    src_row = data_start_row + m

    # A: Date
    ws2.cell(row=r, column=1).value = f"='Monthly Data'!A{src_row}"
    style_data_cell(ws2.cell(row=r, column=1))

    # B: Core Growth
    ws2.cell(row=r, column=2).value = f"='Monthly Data'!K{src_row}"
    style_data_cell(ws2.cell(row=r, column=2), num_fmt)

    # C: Core Peak
    if m == 0:
        ws2.cell(row=r, column=3).value = f"=B{r}"
    else:
        ws2.cell(row=r, column=3).value = f"=MAX(C{r-1},B{r})"
    style_data_cell(ws2.cell(row=r, column=3), num_fmt)

    # D: Core Drawdown = (Peak - Growth) / Peak
    ws2.cell(row=r, column=4).value = f"=(C{r}-B{r})/C{r}"
    style_data_cell(ws2.cell(row=r, column=4), pct_fmt)

    # E: Stacked Growth
    ws2.cell(row=r, column=5).value = f"='Monthly Data'!L{src_row}"
    style_data_cell(ws2.cell(row=r, column=5), num_fmt)

    # F: Stacked Peak
    if m == 0:
        ws2.cell(row=r, column=6).value = f"=E{r}"
    else:
        ws2.cell(row=r, column=6).value = f"=MAX(F{r-1},E{r})"
    style_data_cell(ws2.cell(row=r, column=6), num_fmt)

    # G: Stacked Drawdown
    ws2.cell(row=r, column=7).value = f"=(F{r}-E{r})/F{r}"
    style_data_cell(ws2.cell(row=r, column=7), pct_fmt)

dd_last = 4 + num_months

# Summary at bottom
ws2.cell(row=dd_last + 2, column=1, value="Max Drawdown").font = header_font
ws2.cell(row=dd_last + 2, column=2, value="Core").font = header_font
ws2.cell(row=dd_last + 2, column=3, value="Stacked").font = header_font

ws2.cell(row=dd_last + 3, column=1, value="Max Drawdown")
ws2.cell(row=dd_last + 3, column=1).font = Font(name="DM Sans", bold=True)
ws2.cell(row=dd_last + 3, column=2).value = f"=MAX(D4:D{dd_last})"
style_data_cell(ws2.cell(row=dd_last + 3, column=2), pct_fmt)
ws2.cell(row=dd_last + 3, column=3).value = f"=MAX(G4:G{dd_last})"
style_data_cell(ws2.cell(row=dd_last + 3, column=3), pct_fmt)

# Also put Calmar ratio
ws2.cell(row=dd_last + 4, column=1, value="Calmar Ratio")
ws2.cell(row=dd_last + 4, column=1).font = Font(name="DM Sans", bold=True)
# Reference annualized return from Monthly Data stats
ann_ret_row = sr + 1 + 2  # row of "Annualized Return" in stats
ws2.cell(row=dd_last + 4, column=2).value = f"='Monthly Data'!B{ann_ret_row}/B{dd_last+3}"
style_data_cell(ws2.cell(row=dd_last + 4, column=2), '0.0000')
ws2.cell(row=dd_last + 4, column=3).value = f"='Monthly Data'!C{ann_ret_row}/C{dd_last+3}"
style_data_cell(ws2.cell(row=dd_last + 4, column=3), '0.0000')


# ════════════════════════════════════════
# SHEET 3: Calendar Year Returns
# ════════════════════════════════════════
ws3 = wb.create_sheet("Calendar Year Returns")
ws3.cell(row=1, column=1, value="Calendar Year Performance").font = title_font

headers3 = ["Year", "Core Return\n(100% ACWI)", "Stacked Return\n(ACWI + MF)", "Difference\n(Stacked - Core)", "Overlay Contribution\n(MF - T-Bill)"]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=c, value=h)
style_header(ws3, 3, len(headers3))

for c in range(1, 6):
    ws3.column_dimensions[get_column_letter(c)].width = 20

# Build year data from the monthly returns
# We need to compute these from the raw data since PRODUCT-based formulas
# over filtered ranges are complex. We'll use a helper approach:
# put each month's year in a helper column on Monthly Data, then use
# SUMPRODUCT or manual grouping here.

# Simpler: compute year boundaries and reference growth values
# Year return = Growth(Dec) / Growth(Dec prior year) - 1

# Collect year boundaries from dates
year_rows = {}  # year -> list of data rows
for m in range(num_months):
    idx = common_start + m + 1
    yr = int(dates[idx][:4])
    data_row = data_start_row + m + 1
    if yr not in year_rows:
        year_rows[yr] = []
    year_rows[yr].append(data_row)

years_sorted = sorted(year_rows.keys())

# For each year, compute return as product of (1+monthly return) - 1
# Using PRODUCT formula over the monthly return cells for that year
yr_row = 4
for yr in years_sorted:
    rows = year_rows[yr]
    first_r = rows[0]
    last_r = rows[-1]

    ws3.cell(row=yr_row, column=1, value=yr)
    ws3.cell(row=yr_row, column=1).font = Font(name="DM Sans", bold=True)
    style_data_cell(ws3.cell(row=yr_row, column=1), '0')

    # Core year return: PRODUCT(1+H_first:H_last) - 1
    # We need individual (1+H) references since PRODUCT doesn't do arithmetic
    # Use: product of (1 + each monthly return) - 1
    # Excel doesn't support PRODUCT(1+range) directly, so we use MMULT or a workaround
    # Actually, we can use: EXP(SUMPRODUCT(LN(1+range))) - 1 which is equivalent
    ws3.cell(row=yr_row, column=2).value = f"=EXP(SUMPRODUCT(LN(1+'Monthly Data'!H{first_r}:'Monthly Data'!H{last_r})))-1"
    style_data_cell(ws3.cell(row=yr_row, column=2), pct_fmt)

    # Stacked year return
    ws3.cell(row=yr_row, column=3).value = f"=EXP(SUMPRODUCT(LN(1+'Monthly Data'!J{first_r}:'Monthly Data'!J{last_r})))-1"
    style_data_cell(ws3.cell(row=yr_row, column=3), pct_fmt)

    # Difference
    ws3.cell(row=yr_row, column=4).value = f"=C{yr_row}-B{yr_row}"
    style_data_cell(ws3.cell(row=yr_row, column=4), pct_fmt)

    # Overlay contribution (MF - T-Bill compounded)
    ws3.cell(row=yr_row, column=5).value = f"=EXP(SUMPRODUCT(LN(1+'Monthly Data'!I{first_r}:'Monthly Data'!I{last_r})))-1"
    style_data_cell(ws3.cell(row=yr_row, column=5), pct_fmt)

    yr_row += 1

# Totals row
yr_row += 1
ws3.cell(row=yr_row, column=1, value="Full Period").font = Font(name="DM Sans", bold=True, size=11)
ws3.cell(row=yr_row, column=1).border = thin_border

# Full period = product of all years... or just reference the cumulative from Monthly Data
ws3.cell(row=yr_row, column=2).value = f"='Monthly Data'!K{last_data_row}/'Monthly Data'!K{data_start_row}-1"
style_data_cell(ws3.cell(row=yr_row, column=2), pct_fmt)
ws3.cell(row=yr_row, column=3).value = f"='Monthly Data'!L{last_data_row}/'Monthly Data'!L{data_start_row}-1"
style_data_cell(ws3.cell(row=yr_row, column=3), pct_fmt)
ws3.cell(row=yr_row, column=4).value = f"=C{yr_row}-B{yr_row}"
style_data_cell(ws3.cell(row=yr_row, column=4), pct_fmt)

# Averages
yr_row += 1
ws3.cell(row=yr_row, column=1, value="Average Year").font = Font(name="DM Sans", bold=True, size=11)
ws3.cell(row=yr_row, column=1).border = thin_border
ws3.cell(row=yr_row, column=2).value = f"=AVERAGE(B4:B{yr_row-2})"
style_data_cell(ws3.cell(row=yr_row, column=2), pct_fmt)
ws3.cell(row=yr_row, column=3).value = f"=AVERAGE(C4:C{yr_row-2})"
style_data_cell(ws3.cell(row=yr_row, column=3), pct_fmt)
ws3.cell(row=yr_row, column=4).value = f"=C{yr_row}-B{yr_row}"
style_data_cell(ws3.cell(row=yr_row, column=4), pct_fmt)

# Best/Worst
yr_row += 1
ws3.cell(row=yr_row, column=1, value="Best Year").font = Font(name="DM Sans", bold=True, size=11)
ws3.cell(row=yr_row, column=1).border = thin_border
ws3.cell(row=yr_row, column=2).value = f"=MAX(B4:B{yr_row-3})"
style_data_cell(ws3.cell(row=yr_row, column=2), pct_fmt)
ws3.cell(row=yr_row, column=3).value = f"=MAX(C4:C{yr_row-3})"
style_data_cell(ws3.cell(row=yr_row, column=3), pct_fmt)

yr_row += 1
ws3.cell(row=yr_row, column=1, value="Worst Year").font = Font(name="DM Sans", bold=True, size=11)
ws3.cell(row=yr_row, column=1).border = thin_border
ws3.cell(row=yr_row, column=2).value = f"=MIN(B4:B{yr_row-4})"
style_data_cell(ws3.cell(row=yr_row, column=2), pct_fmt)
ws3.cell(row=yr_row, column=3).value = f"=MIN(C4:C{yr_row-4})"
style_data_cell(ws3.cell(row=yr_row, column=3), pct_fmt)

# Win rate
yr_row += 1
ws3.cell(row=yr_row, column=1, value="% Years Stacked > Core").font = Font(name="DM Sans", bold=True, size=11)
ws3.cell(row=yr_row, column=1).border = thin_border
ws3.cell(row=yr_row, column=3).value = f"=COUNTIF(D4:D{yr_row-5},\">0\")/COUNT(D4:D{yr_row-5})"
style_data_cell(ws3.cell(row=yr_row, column=3), pct_fmt)


# ════════════════════════════════════════
# SHEET 4: Methodology Notes
# ════════════════════════════════════════
ws4 = wb.create_sheet("Methodology")
ws4.column_dimensions["A"].width = 80

notes = [
    "WIDGET CALCULATION METHODOLOGY (from build_widget.py computePortfolio())",
    "",
    "1. MONTHLY RETURNS",
    "   Each asset's monthly return = currentIndexLevel / previousIndexLevel - 1",
    "",
    "2. CORE RETURN (per month)",
    "   coreReturn = SUM(weight_i * assetReturn_i) for all core assets",
    "   In this QC: coreReturn = 1.0 * ACWI_return (single asset, 100% weight)",
    "",
    "3. STACKED RETURN (per month)",
    "   stackedReturn starts as a copy of coreReturn, then for each overlay asset:",
    "     stackedReturn += weight * overlayReturn",
    "     stackedReturn -= weight * tBillReturn     (financing cost)",
    "     stackedReturn -= weight * (feeBp/10000/12 + financingBp/10000/12)",
    "",
    "   In this QC (all fees = 0):",
    "     stackedReturn = coreReturn + (1.0 * MF_return) - (1.0 * TBill_return)",
    "     stackedReturn = ACWI_return + MF_return - TBill_return",
    "",
    "4. GROWTH OF $1",
    "   growth[0] = 1",
    "   growth[i] = growth[i-1] * (1 + monthlyReturn - monthlyAdvisoryFee)",
    "   Advisory fee = 0 here, so: growth[i] = growth[i-1] * (1 + monthlyReturn)",
    "",
    "5. STATISTICS (from computeStats())",
    "   - Cumulative Return: product(1 + r_i) - 1",
    "   - Annualized Return: (1 + cumReturn)^(12/N) - 1",
    "   - Volatility: STDEV(monthly returns) * SQRT(12)",
    "   - Sharpe: (annReturn - rfAnnualized) / vol",
    "     where rfAnnualized = average monthly T-Bill return * 12",
    "   - Sortino: annReturn / downsideDev",
    "     where downsideDev = SQRT(sum(min(r,0)^2) / N) * SQRT(12)",
    "   - Max DD: max of (peak - value) / peak over growth series",
    "   - Calmar: annReturn / maxDD",
    "   - Skewness: Excel SKEW equivalent (sample-adjusted)",
    "   - Kurtosis: Excel KURT equivalent (excess, sample-adjusted)",
    "",
    "6. CALENDAR YEAR RETURNS",
    "   yearReturn = product(1 + monthly_r for months in year) - 1",
    "",
    "7. KEY ASSUMPTION",
    "   The overlay 'borrows' at the T-Bill rate. The net overlay contribution is",
    "   the excess return of managed futures ABOVE T-Bills. This is the standard",
    "   return-stacking financing model.",
]

for i, line in enumerate(notes, 1):
    ws4.cell(row=i, column=1, value=line).font = Font(name="Consolas", size=10)

# ── Save ──
out_path = "QC_Stacked_Portfolio_ACWI_MFT.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Period: {dates[common_start]} to {dates[common_end]} ({num_months} months)")
print(f"Sheets: Monthly Data, Drawdown, Calendar Year Returns, Methodology")
