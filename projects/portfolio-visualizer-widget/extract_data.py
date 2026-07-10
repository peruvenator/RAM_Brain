"""
Extract all data from Return_Stacking_Visualizer.xlsx for the HTML widget.
Outputs JSON files to data/ directory.
"""
import json
import os
import openpyxl
from datetime import datetime

EXCEL_PATH = "Return_Stacking_Visualizer.xlsx"
OUTPUT_DIR = "data"

os.makedirs(OUTPUT_DIR, exist_ok=True)

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)


# --- 1. Index Map ---
# Maps asset short names to Bloomberg index names, start dates, and categories
ws = wb["Index Map"]
index_map = []
for row in ws.iter_rows(min_row=2, max_row=65, values_only=False):
    short_name = row[0].value  # col A
    bloomberg_name = row[1].value  # col B
    start_date = row[3].value  # col D
    asset_class = row[4].value  # col E
    if short_name and bloomberg_name:
        # Categorize custom assets properly
        if asset_class and asset_class != 0:
            cat = asset_class
        elif "Custom Asset" in short_name:
            cat = "Custom"
        else:
            cat = "Unknown"
        entry = {
            "shortName": short_name,
            "bloombergName": bloomberg_name,
            "startDate": start_date.strftime("%Y-%m-%d") if isinstance(start_date, datetime) else str(start_date) if start_date else None,
            "assetClass": cat,
        }
        index_map.append(entry)

with open(os.path.join(OUTPUT_DIR, "index_map.json"), "w") as f:
    json.dump(index_map, f, indent=2)
print(f"Index Map: {len(index_map)} assets extracted")


# --- 2. Indices Monthly ---
# All index time series as price levels
ws = wb["Indices Monthly"]

# Get column headers (row 1)
headers = {}
for cell in ws[1]:
    if cell.value and not str(cell.value).startswith("="):
        headers[cell.column] = cell.value

# Find the date column (C)
date_col = None
data_cols = {}
for col_idx, name in headers.items():
    if name == "Date":
        date_col = col_idx
    else:
        data_cols[col_idx] = name

print(f"Indices Monthly: {len(data_cols)} index columns, date col={date_col}")

# Extract all data rows
indices_data = {"dates": [], "series": {name: [] for name in data_cols.values()}}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    date_cell = row[date_col - 1] if date_col else None
    if date_cell and date_cell.value and isinstance(date_cell.value, datetime):
        date_str = date_cell.value.strftime("%Y-%m-%d")
        indices_data["dates"].append(date_str)
        for col_idx, name in data_cols.items():
            val = row[col_idx - 1].value
            if val is not None and val != "" and val != "#N/A":
                try:
                    indices_data["series"][name].append(float(val))
                except (ValueError, TypeError):
                    indices_data["series"][name].append(None)
            else:
                indices_data["series"][name].append(None)

# Remove series that are entirely null
empty_series = [k for k, v in indices_data["series"].items() if all(x is None for x in v)]
for k in empty_series:
    del indices_data["series"][k]
    print(f"  Removed empty series: {k}")

print(f"  {len(indices_data['dates'])} dates, {len(indices_data['series'])} non-empty series")

with open(os.path.join(OUTPUT_DIR, "indices_monthly.json"), "w") as f:
    json.dump(indices_data, f)
print(f"  Saved to indices_monthly.json ({os.path.getsize(os.path.join(OUTPUT_DIR, 'indices_monthly.json')) / 1024:.0f} KB)")


# --- 3. Custom Asset Returns ---
ws = wb["Custom Asset Returns"]
custom_assets = {"dates": [], "series": {}}

# Headers in row 2
custom_headers = {}
for cell in ws[2]:
    if cell.value and cell.column >= 2:  # col B onwards
        custom_headers[cell.column] = cell.value

for name in custom_headers.values():
    custom_assets["series"][name] = []

for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
    date_val = row[0].value  # col A
    if date_val and isinstance(date_val, datetime):
        custom_assets["dates"].append(date_val.strftime("%Y-%m-%d"))
        for col_idx, name in custom_headers.items():
            val = row[col_idx - 1].value
            if val is not None and val != "":
                try:
                    custom_assets["series"][name].append(float(val))
                except (ValueError, TypeError):
                    custom_assets["series"][name].append(None)
            else:
                custom_assets["series"][name].append(None)

# Only keep custom assets that have actual data
for name in list(custom_assets["series"].keys()):
    if all(x is None for x in custom_assets["series"][name]):
        del custom_assets["series"][name]

print(f"Custom Assets: {len(custom_assets['series'])} with data, {len(custom_assets['dates'])} dates")

with open(os.path.join(OUTPUT_DIR, "custom_assets.json"), "w") as f:
    json.dump(custom_assets, f)


# --- 4. Disclosures ---
ws = wb["Disclosures"]
disclosures = {"general": [], "indexDefinitions": []}

in_definitions = False
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    # Content is in column B
    val = row[1].value if len(row) > 1 else None
    if val:
        val = str(val).strip()
        if val == "Index Definitions":
            in_definitions = True
            continue
        if val == "IMPORTANT DISCLOSURES":
            continue
        if in_definitions:
            disclosures["indexDefinitions"].append(val)
        else:
            disclosures["general"].append(val)

print(f"Disclosures: {len(disclosures['general'])} general paragraphs, {len(disclosures['indexDefinitions'])} index definitions")

with open(os.path.join(OUTPUT_DIR, "disclosures.json"), "w") as f:
    json.dump(disclosures, f, indent=2)


# --- 5. Asset class groupings from Index Map ---
# Build the grouped asset menu structure
categories = {}
for asset in index_map:
    cat = asset["assetClass"]
    if cat not in categories:
        categories[cat] = []
    categories[cat].append(asset["shortName"])

print("\nAsset class categories:")
for cat, assets in categories.items():
    print(f"  {cat}: {len(assets)} assets")


# --- 6. Validation: Check a known portfolio return ---
# Portfolio 1 in Excel: 60% US Large Cap + 40% US Core Fixed Income
# Expected cumulative return ~4.36x (from Portfolio 1 sheet, cell C3)
print("\n--- Validation ---")
us_large_cap_idx = None
us_bonds_idx = None
for asset in index_map:
    if asset["shortName"] == "U.S. Large Cap Equities":
        us_large_cap_idx = asset["bloombergName"]
    if asset["shortName"] == "U.S. Core Fixed Income":
        us_bonds_idx = asset["bloombergName"]

print(f"US Large Cap -> {us_large_cap_idx}")
print(f"US Core Fixed Income -> {us_bonds_idx}")

if us_large_cap_idx and us_bonds_idx:
    stocks = indices_data["series"].get(us_large_cap_idx, [])
    bonds = indices_data["series"].get(us_bonds_idx, [])
    dates = indices_data["dates"]

    # Find start date (12/31/1999) index
    start_idx = None
    for i, d in enumerate(dates):
        if d >= "1999-12-31":
            start_idx = i
            break

    if start_idx is not None and stocks[start_idx] and bonds[start_idx]:
        # Calculate cumulative return from start to end
        end_idx = len(dates) - 1
        # Find last non-None values
        while stocks[end_idx] is None or bonds[end_idx] is None:
            end_idx -= 1

        stock_return = stocks[end_idx] / stocks[start_idx]
        bond_return = bonds[end_idx] / bonds[start_idx]
        portfolio_return = 0.6 * stock_return + 0.4 * bond_return

        print(f"Period: {dates[start_idx]} to {dates[end_idx]}")
        print(f"Stock cumulative: {stock_return:.4f}x")
        print(f"Bond cumulative: {bond_return:.4f}x")
        print(f"60/40 portfolio: {portfolio_return:.4f}x")
        print(f"Expected from Excel: ~4.36x (cumulative return)")
        print(f"Note: Excel cumulative return = ending value - 1, so ending value = 5.36x")


print("\nDone! All data extracted to data/ directory.")
