"""
Build Excel file mapping widget names to source index names and Bloomberg tickers.
Columns: Widget Name | Full Index Name (from disclosures) | Bloomberg Ticker | Notes
"""

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── Styles ──
header_font_white = Font(name="DM Sans", bold=True, size=10, color="FFFFFF")
header_fill = PatternFill(start_color="323A46", end_color="323A46", fill_type="solid")
title_font = Font(name="DM Sans", bold=True, size=14)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
note_font = Font(name="DM Sans", size=9, italic=True, color="666666")

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

# ── Complete mapping: (widget_name, full_index_name, bloomberg_ticker, notes) ──
# Parsed from disclosures + web research
mapping = [
    # ── Equity ──
    ("U.S. Large Cap Equities", "MSCI USA (USD) Total Return Index",
     "NDDUUS", "MSCI Net Total Return USD; verify on BBG terminal"),
    ("U.S. Value", "MSCI USA Value (USD) Total Return Index",
     "M1USVA", "MSCI factor index; verify TR variant on BBG"),
    ("U.S. Growth", "MSCI USA Growth (USD) Total Return Index",
     "M1USGR", "MSCI factor index; verify TR variant on BBG"),
    ("U.S. Momentum", "MSCI USA Momentum (USD) Total Return Index",
     "M1USMO", "MSCI factor index; verify TR variant on BBG"),
    ("U.S. Quality", "MSCI USA Quality (USD) Total Return Index",
     "M1USQU", "MSCI factor index; verify TR variant on BBG"),
    ("U.S. Low Volatility", "MSCI USA Minimum Volatility (USD) Total Return Index",
     "M1USMVOL", "MSCI factor index; verify TR variant on BBG"),
    ("U.S. Small Cap Equities", "Russell 2000 Total Return Index",
     "RU20INTR", ""),
    ("U.S. Small Cap Value", "Russell 2000 Value Total Return Index",
     "RU20VATR", ""),
    ("International Equities", "MSCI EAFE Net Return Index USD",
     "NDDUEAFE", "MSCI Net Total Return USD"),
    ("International Value", "MSCI EAFE Value Net Return Index USD",
     "NDDUEAFV", "MSCI Net TR; verify on BBG"),
    ("Global Equities", "MSCI ACWI Net Return Index USD",
     "NDUEACWF", "Confirmed"),
    ("Emerging Market Equities", "MSCI Emerging Markets Net Return Index USD",
     "NDUEEGF", "MSCI Net TR USD"),
    ("Emerging Markets Value", "MSCI Emerging Markets Value Net Return Index USD",
     "NDUEVEMF", "MSCI Net TR; verify on BBG"),

    # ── Fixed Income ──
    ("U.S. Treasury Ladder", "ReSolve U.S. Treasury Ladder Index",
     "N/A", "Proprietary ReSolve index; no public BBG ticker"),
    ("U.S. Core Fixed Income", "Bloomberg US Agg Total Return Value Unhedged USD",
     "LBUSTRUU", "Confirmed"),
    ("U.S. Treasuries", "Bloomberg US Treasury Total Return Index",
     "LUATTRUU", "Confirmed"),
    ("Short-Term U.S. Treasuries", "Bloomberg US Treasury 1-5 Year Total Return Index",
     "LTR1TRUU", "Confirmed"),
    ("Intermediate-Term U.S. Treasuries", "Bloomberg US Intermediate Treasury TR Index Value Unhedged USD",
     "LT08TRUU", "Confirmed"),
    ("Long-Term U.S. Treasuries", "Bloomberg US Long Treasury Total Return Index Value Unhedged USD",
     "LUTLTRUU", "Confirmed"),
    ("U.S. Corporate Fixed Income", "Bloomberg US Corporate Total Return Value Unhedged USD",
     "LUACTRUU", "Confirmed"),
    ("Short-Term Investment Grade Fixed Income", "Bloomberg US Credit 1-5 Years Total Return Index Value Unhedged USD",
     "LDC5TRUU", "Confirmed"),
    ("Intermediate-Term Investment Grade Fixed Income", "Bloomberg US Credit Corp 5-10Y Total Return Index Value Unhedged USD",
     "BCR5TRUU", "Confirmed"),
    ("Long-Term Investment Grade Fixed Income", "Bloomberg U.S. 10+ Year Corporate Bond Total Return Index Unhedged USD",
     "I13284US", "Confirmed"),
    ("High Yield Fixed Income", "Bloomberg US Corporate High Yield Total Return Index Value Unhedged USD",
     "LF98TRUU", "Confirmed"),
    ("Mortgage Backed Securities", "Bloomberg US MBS Index Total Return Value Unhedged USD",
     "LUMSTRUU", "Confirmed"),
    ("International Treasuries", "Bloomberg Global Treasury ex-US Capped Total Return Index Value Unhedged USD",
     "LGT1TRUU", "Closest match; capped variant may differ -- verify on BBG"),
    ("Global Core Fixed Income", "Bloomberg Global-Aggregate Total Return Index Value Unhedged USD",
     "LEGATRUU", "Confirmed"),
    ("Emerging Market Bonds (Local)", "JPMorgan GBI-EM Global Diversified (Local Currency)",
     "JGENVUUG", "JPM GBI-EM Global Diversified Unhedged; verify on BBG"),
    ("Emerging Market Bonds (USD)", "JPMorgan EMBI Global Diversified (USD)",
     "JPGCCOMP", "JPM EMBI Global Diversified Composite; verify on BBG"),

    # ── Real Assets ──
    ("Real Estate", "Dow Jones U.S. Select REIT Total Return Index",
     "DWRTFT", "Confirmed (DWRTF = price return)"),
    ("Commodities", "S&P GSCI Index (Spot)",
     "SPGSCI", "Confirmed (spot); SPGSCITR = total return"),
    ("Gold", "XAU Currency (Gold Spot USD)",
     "XAU Curncy", "Confirmed"),
    ("Digital Assets", "S&P Bitcoin Index",
     "SPBTC", "Confirmed"),
    ("Short-Term TIPS", "Bloomberg US Treasury TIPS 0-5 Years Total Return Index Unhedged USD",
     "LTP5TRUU", "Confirmed"),
    ("Intermediate-Term TIPS", "Bloomberg US Treasury Inflation Notes TR Index Value Unhedged USD",
     "LBUTTRUU", "All maturities TIPS TR; verify on BBG"),

    # ── Alternatives (Hedge Fund) ──
    ("Long Volatility", "CBOE Eurekahedge Long Volatility Hedge Fund Index",
     "EHFI451", "Confirmed"),
    ("Tail Risk Hedges", "CBOE Eurekahedge Tail Risk Hedge Fund Index",
     "EHFI453", "Confirmed"),
    ("Equity Long/Short", "PivotalPath Equity Diversified: Global Long/Short Index",
     "EQDGLS", "PivotalPath internal ticker; available via BBG PIVT page"),
    ("Hedge Funds", "PivotalPath Composite EW Index",
     "HFCEW", "PivotalPath internal ticker; available via BBG PIVT page"),
    ("Equity Market Neutral", "EurekaHedge Equity Market Neutral Hedge Fund Index",
     "EHFI751", "Confirmed"),
    ("Event Driven", "PivotalPath Event Driven Index",
     "EVD", "PivotalPath internal ticker"),
    ("Merger Arbitrage", "PivotalPath Event Driven: Merger Arbitrage Index",
     "EVDMER", "PivotalPath internal ticker"),
    ("Relative Value", "PivotalPath Credit: Relative Value Index",
     "CRDREL", "PivotalPath internal ticker"),
    ("Managed Futures CTA", "PivotalPath Managed Futures Index",
     "MFT", "PivotalPath internal ticker"),
    ("Managed Futures Trend", "SG Trend Index",
     "NEIXCTAT", "Societe Generale; confirmed"),
    ("Risk Parity (10%)", "S&P Risk Parity Index - 10% Target Volatility",
     "SPRP10UT", "S&P Dow Jones; verify on BBG"),
    ("Risk Parity (12%)", "S&P Risk Parity Index - 12% Target Volatility",
     "SPRP12UT", "S&P Dow Jones; verify on BBG"),
    ("Risk Parity (15%)", "S&P Risk Parity Index - 15% Target Volatility",
     "SPRP15UT", "S&P Dow Jones; verify on BBG"),
    ("Global Macro", "PivotalPath Global Macro Index",
     "GBM", "PivotalPath internal ticker"),
    ("Global Macro (Commodities)", "PivotalPath Global Macro: Commodities Index",
     "GBMCOM", "PivotalPath internal ticker"),
    ("Systematic Global Macro", "PivotalPath Global Macro: Quantitative Index",
     "GBMQNT", "PivotalPath internal ticker"),
    ("Global Macro (Risk Premia)", "PivotalPath Global Macro: Risk Premia Index",
     "GBMRPM", "PivotalPath internal ticker"),
    ("Global Stock / Bond Momentum", "Newfound/ReSolve Robust Equity Momentum TR Index",
     "NRROMOT", "Confirmed; calculated by Solactive AG"),
    ("Risk-Weighted Gold/Bitcoin", "Kaiko ByteTree BOLD1 Inverse Volatility Index",
     "BOLD1", "ByteTree; available via BBG"),

    # ── Cash ──
    ("Cash", "Bloomberg US Treasury Bill Total Return Index Unhedged USD",
     "LD12TRUU", "Verify on BBG; may also be LB1MTRUU"),

    # ── Other ──
    ("Futures Yield (Carry)", "Futures Yield (Carry)",
     "N/A", "Proprietary/custom series; no public BBG ticker"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Index Ticker Map"

ws.cell(row=1, column=1, value="Widget Asset Name to Index Ticker Mapping").font = title_font
ws.cell(row=2, column=1, value="Tickers marked 'verify' should be confirmed on Bloomberg Terminal. PivotalPath tickers are internal identifiers accessible via BBG PIVT <GO>.").font = Font(name="DM Sans", size=9, italic=True)

headers = ["Widget Display Name", "Full Index / Source Name", "Bloomberg Ticker", "Asset Class", "Notes"]
HEADER_ROW = 4
for c, h in enumerate(headers, 1):
    ws.cell(row=HEADER_ROW, column=c, value=h)
style_header(ws, HEADER_ROW, len(headers))

ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 65
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 50

# Load index_map for asset class info
with open("data/index_map.json", "r") as f:
    index_map = json.load(f)

asset_class_lookup = {e["shortName"]: e["assetClass"] for e in index_map}

for i, (widget_name, full_name, ticker, notes) in enumerate(mapping):
    row = HEADER_ROW + 1 + i
    ws.cell(row=row, column=1, value=widget_name)
    ws.cell(row=row, column=1).font = Font(name="DM Sans", bold=True)
    ws.cell(row=row, column=2, value=full_name)
    ws.cell(row=row, column=3, value=ticker)
    ws.cell(row=row, column=3).font = Font(name="DM Sans", bold=True)
    ws.cell(row=row, column=4, value=asset_class_lookup.get(widget_name, ""))
    ws.cell(row=row, column=5, value=notes)
    ws.cell(row=row, column=5).font = note_font

    for c in range(1, 6):
        ws.cell(row=row, column=c).border = thin_border
        if c != 1 and c != 2 and c != 5:
            ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")

    # Color code: green for confirmed, yellow for verify
    if "Confirmed" in notes or notes == "":
        ws.cell(row=row, column=3).font = Font(name="DM Sans", bold=True, color="006600")
    elif "verify" in notes.lower():
        ws.cell(row=row, column=3).font = Font(name="DM Sans", bold=True, color="996600")
    elif "N/A" in ticker or "Proprietary" in notes:
        ws.cell(row=row, column=3).font = Font(name="DM Sans", bold=True, color="999999")

out_path = "Index_Ticker_Map.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Total entries: {len(mapping)}")
