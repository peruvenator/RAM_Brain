"""Build Widget_AssetClass_Data_MFMAGold.xlsx mirroring Widget_AssetClass_Data.xlsx.

Overlay is a 33.34% Managed Futures CTA / 33.33% Merger Arbitrage / 33.33% Gold
equal-weight blend (monthly rebalanced). Stacked columns put that blend on top
of U.S. Large Cap and International Equity cores, financed at T-Bills.
"""
import json
import openpyxl
from openpyxl.styles import Font

OUT = "Widget_AssetClass_Data_MFMAGold.xlsx"
BASE_DATE = "1999-12-31"

WEIGHTS = {"MFT": 0.3334, "EVDMER": 0.3333, "Gold (Spot)": 0.3333}

with open("data/indices_compact.json") as f:
    d = json.load(f)

dates = d["dates"]
series = d["series"]
base_idx = dates.index(BASE_DATE)

def levels(key):
    s = series[key]
    return [None] * s["start"] + s["values"]

raw = {k: levels(k) for k in ["MFT", "EVDMER", "Gold (Spot)", "Treasury Bill", "MSCI USA", "MSCI EAFE"]}

dates_out = dates[base_idx:]
n = len(dates_out)

def rebase(key):
    base = raw[key][base_idx]
    return [raw[key][base_idx + i] / base * 100 for i in range(n)]

mft = rebase("MFT")
ma = rebase("EVDMER")
gold = rebase("Gold (Spot)")
tbill = rebase("Treasury Bill")
usa = rebase("MSCI USA")
eafe = rebase("MSCI EAFE")

# Equal-weight blend, monthly rebalanced, base 100
blend = [100.0]
for i in range(1, n):
    r = (WEIGHTS["MFT"] * (mft[i] / mft[i-1] - 1)
         + WEIGHTS["EVDMER"] * (ma[i] / ma[i-1] - 1)
         + WEIGHTS["Gold (Spot)"] * (gold[i] / gold[i-1] - 1))
    blend.append(blend[-1] * (1 + r))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Index Levels"

headers1 = [
    "Date",
    "Managed Futures CTA",
    "Merger Arbitrage",
    "Gold",
    "Equal-Weight Blend (33.34% MFCTA / 33.33% MA / 33.33% Gold)",
    "U.S. Large Cap Equities",
    "International Equities",
    "Risk-Free (T-Bills)",
    "U.S. Large Cap + Equal-Weight Blend - Cash",
    "International Equities + Equal-Weight Blend - Cash",
]
headers2 = [
    "Underlying index",
    "MFT",
    "EVDMER (Merger Arbitrage)",
    "Gold (Spot)",
    "33.34% MFT + 33.33% EVDMER + 33.33% Gold (Spot)",
    "MSCI USA",
    "MSCI EAFE",
    "Bloomberg US T-Bill TR (Treasury Bill)",
    "MSCI USA + Blend - T-Bills",
    "MSCI EAFE + Blend - T-Bills",
]

bold = Font(bold=True)
for c, h in enumerate(headers1, 1):
    cell = ws.cell(1, c, h); cell.font = bold
for c, h in enumerate(headers2, 1):
    ws.cell(2, c, h)

# Row 3 = base date, all rebased columns at 100; stacked columns at 100
ws.cell(3, 1, dates_out[0])
for c in range(2, 9):
    ws.cell(3, c, 100)
ws.cell(3, 9, 100)
ws.cell(3, 10, 100)

cols_data = [None, mft, ma, gold, blend, usa, eafe, tbill]

for i in range(1, n):
    r = 3 + i
    ws.cell(r, 1, dates_out[i])
    for c in range(2, 9):
        ws.cell(r, c, cols_data[c-1][i])
    # Stacked = prev * (1 + r_equity + r_blend - r_tbill)
    # I col: MSCI USA (col F) + Blend (col E) - T-Bills (col H)
    ws.cell(r, 9, f"=I{r-1}*(1+(F{r}/F{r-1}-1)+(E{r}/E{r-1}-1)-(H{r}/H{r-1}-1))")
    # J col: MSCI EAFE (col G) + Blend (col E) - T-Bills (col H)
    ws.cell(r, 10, f"=J{r-1}*(1+(G{r}/G{r-1}-1)+(E{r}/E{r-1}-1)-(H{r}/H{r-1}-1))")

# Column widths
widths = [12, 22, 22, 16, 50, 24, 24, 22, 42, 46]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Monthly Returns sheet
ws2 = wb.create_sheet("Monthly Returns")
for c, h in enumerate(headers1, 1):
    cell = ws2.cell(1, c, h); cell.font = bold
for c, h in enumerate(headers2, 1):
    ws2.cell(2, c, h)

for i in range(1, n):
    r = 2 + i
    ws2.cell(r, 1, dates_out[i])
    ws2.cell(r, 2, mft[i] / mft[i-1] - 1)
    ws2.cell(r, 3, ma[i] / ma[i-1] - 1)
    ws2.cell(r, 4, gold[i] / gold[i-1] - 1)
    ws2.cell(r, 5, blend[i] / blend[i-1] - 1)
    ws2.cell(r, 6, usa[i] / usa[i-1] - 1)
    ws2.cell(r, 7, eafe[i] / eafe[i-1] - 1)
    ws2.cell(r, 8, tbill[i] / tbill[i-1] - 1)
    # Stacked monthly return = r_equity + r_blend - r_tbill
    ws2.cell(r, 9, f"=F{r}+E{r}-H{r}")
    ws2.cell(r, 10, f"=G{r}+E{r}-H{r}")

for i, w in enumerate(widths, 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Notes sheet
ws3 = wb.create_sheet("Notes")
notes = [
    f"Columns added (base 100 at {BASE_DATE} for index columns):",
    "  - Risk-Free (T-Bills): Bloomberg US Treasury Bill Total Return Index, rebased to 100.",
    "  - Equal-Weight Blend: 33.34% Managed Futures CTA (MFT) + 33.33% Merger Arbitrage (EVDMER) + 33.33% Gold (Spot), monthly rebalanced, rebased to 100.",
    "  - Stacked columns are monthly return-stacked series, compounded to a growth index from 100.",
    "",
    "Stacked return formula (monthly): r = r(equity) + ( r(overlay) - r(T-Bills) )",
    "  - Equity leg is held 100% unlevered (the core).",
    "  - Overlay (Equal-Weight Blend) is held 100% notional, financed at T-Bills (cost of cash).",
    "  - Managed Futures CTA, Merger Arbitrage, and Gold are total-return indices, so subtracting T-Bills converts the blend to the financed overlay (excess over cash).",
    "",
    "Stacked columns:",
    "  H  Risk-Free (T-Bills)",
    "  I  U.S. Large Cap (MSCI USA) + Equal-Weight Blend (33.34% MFCTA / 33.33% MA / 33.33% Gold) - Cash",
    "  J  International Equities (MSCI EAFE) + Equal-Weight Blend (33.34% MFCTA / 33.33% MA / 33.33% Gold) - Cash",
    "",
    f"Data span: {dates_out[0]} to {dates_out[-1]} ({n} monthly observations).",
]
for i, txt in enumerate(notes, 1):
    ws3.cell(i, 1, txt)
ws3.column_dimensions["A"].width = 120

wb.save(OUT)
print(f"Wrote {OUT}: {n} level rows, {n-1} monthly return rows ({dates_out[0]} -> {dates_out[-1]})")
