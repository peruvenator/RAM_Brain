"""
Export the two Merger Arbitrage series to Excel:
  - Merger Arbitrage          (source: EVDMER)
  - Merger Arbitrage (AB)     (source: AlphaBeta Merger Arbitrage Index)

Output columns per series: index level + monthly return.
"""

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

with open("data/indices_compact.json", "r") as f:
    data = json.load(f)

dates = data["dates"]
num_dates = len(dates)

SERIES = [
    {"widget_name": "Merger Arbitrage",      "source": "EVDMER"},
    {"widget_name": "Merger Arbitrage (AB)", "source": "AlphaBeta Merger Arbitrage Index"},
]

# Expand each sparse series into a full-length (aligned to dates) level list
for s in SERIES:
    raw = data["series"][s["source"]]
    start = raw["start"]
    vals = raw["values"]
    levels = [None] * num_dates
    for i, v in enumerate(vals):
        levels[start + i] = v
    s["levels"] = levels

# ── Styles ──
header_font_white = Font(name="DM Sans", bold=True, size=10, color="FFFFFF")
header_fill = PatternFill(start_color="323A46", end_color="323A46", fill_type="solid")
title_font = Font(name="DM Sans", bold=True, size=14)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
pct_fmt = '0.00%'
idx_fmt = '#,##0.00'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Merger Arbitrage"
ws.cell(row=1, column=1, value="Merger Arbitrage Series (Index Levels + Monthly Returns)").font = title_font

# Header row
HEADER_ROW = 3
headers = ["Date"]
for s in SERIES:
    headers.append(f"{s['widget_name']} (Level)")
    headers.append(f"{s['widget_name']} (Return)")
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=HEADER_ROW, column=c, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

ws.column_dimensions["A"].width = 12
for c in range(2, len(headers) + 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 22

# Data rows -- only emit rows where at least one series has data
first_data = min(s_idx for s in SERIES for s_idx, v in enumerate(s["levels"]) if v is not None)

out_row = HEADER_ROW + 1
for d_idx in range(first_data, num_dates):
    ws.cell(row=out_row, column=1, value=dates[d_idx]).alignment = Alignment(horizontal="center")
    ws.cell(row=out_row, column=1).border = thin_border
    col = 2
    for s in SERIES:
        lvl = s["levels"][d_idx]
        prev = s["levels"][d_idx - 1] if d_idx > 0 else None
        # level
        lc = ws.cell(row=out_row, column=col)
        lc.border = thin_border
        lc.alignment = Alignment(horizontal="center")
        if lvl is not None:
            lc.value = lvl
            lc.number_format = idx_fmt
        # return
        rc = ws.cell(row=out_row, column=col + 1)
        rc.border = thin_border
        rc.alignment = Alignment(horizontal="center")
        if lvl is not None and prev not in (None, 0):
            rc.value = lvl / prev - 1
            rc.number_format = pct_fmt
        col += 2
    out_row += 1

# Freeze header + date column
ws.freeze_panes = "B4"

out_path = "Merger_Arbitrage_Data.xlsx"
wb.save(out_path)

for s in SERIES:
    present = [d for d, v in zip(dates, s["levels"]) if v is not None]
    print(f"{s['widget_name']}: {len(present)} months, {present[0]} to {present[-1]}")
print(f"\nSaved: {out_path}")
