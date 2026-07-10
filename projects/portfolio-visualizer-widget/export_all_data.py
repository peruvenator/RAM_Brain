"""
Export all widget data series to Excel:
  Tab 1: Raw index levels with original source names as headers
  Tab 2: Widget name <-> original index name mapping
  Tab 3: Calendar year returns for all assets
"""

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── Load data ──
with open("data/indices_compact.json", "r") as f:
    data = json.load(f)

with open("data/index_map.json", "r") as f:
    index_map = json.load(f)

dates = data["dates"]
num_dates = len(dates)

# Build mapping: bloombergName -> shortName, and ordered list
# index_map order matches the series keys order
assets = []
for entry in index_map:
    short = entry["shortName"]
    bloom = entry["bloombergName"]
    asset_class = entry["assetClass"]
    start = entry["startDate"]
    series = data["series"].get(bloom)
    if series:
        assets.append({
            "widget_name": short,
            "source_name": bloom,
            "asset_class": asset_class,
            "start_date": start,
            "series_start": series["start"],
            "values": series["values"],
        })

print(f"Found {len(assets)} assets")

# ── Styles ──
header_font_white = Font(name="DM Sans", bold=True, size=10, color="FFFFFF")
header_fill = PatternFill(start_color="323A46", end_color="323A46", fill_type="solid")
title_font = Font(name="DM Sans", bold=True, size=14)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
pct_fmt = '0.00%'
idx_fmt = '#,##0.00'

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

wb = openpyxl.Workbook()

# ════════════════════════════════════════
# SHEET 1: Raw Index Levels
# ════════════════════════════════════════
ws1 = wb.active
ws1.title = "Index Levels"
ws1.cell(row=1, column=1, value="All Index Levels (Original Source Names)").font = title_font

# Headers: Date + each asset's source name
HEADER_ROW = 3
ws1.cell(row=HEADER_ROW, column=1, value="Date")
for i, a in enumerate(assets):
    ws1.cell(row=HEADER_ROW, column=i + 2, value=a["source_name"])
style_header(ws1, HEADER_ROW, len(assets) + 1)

# Column widths
ws1.column_dimensions["A"].width = 12
for i in range(len(assets)):
    col_letter = openpyxl.utils.get_column_letter(i + 2)
    ws1.column_dimensions[col_letter].width = 14

# Data rows
for d_idx in range(num_dates):
    row = HEADER_ROW + 1 + d_idx
    ws1.cell(row=row, column=1, value=dates[d_idx])
    ws1.cell(row=row, column=1).border = thin_border
    ws1.cell(row=row, column=1).alignment = Alignment(horizontal="center")

    for a_idx, a in enumerate(assets):
        cell = ws1.cell(row=row, column=a_idx + 2)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

        data_i = d_idx - a["series_start"]
        if 0 <= data_i < len(a["values"]) and a["values"][data_i] is not None:
            cell.value = a["values"][data_i]
            cell.number_format = idx_fmt
        # else leave blank

print("Sheet 1 (Index Levels) done")

# ════════════════════════════════════════
# SHEET 2: Name Mapping
# ════════════════════════════════════════
ws2 = wb.create_sheet("Asset Mapping")
ws2.cell(row=1, column=1, value="Widget Asset Name to Source Index Mapping").font = title_font

map_headers = ["Widget Name", "Original Index / Source Name", "Asset Class", "Data Start Date"]
MAP_HEADER_ROW = 3
for c, h in enumerate(map_headers, 1):
    ws2.cell(row=MAP_HEADER_ROW, column=c, value=h)
style_header(ws2, MAP_HEADER_ROW, len(map_headers))

ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 55
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 16

for i, a in enumerate(assets):
    row = MAP_HEADER_ROW + 1 + i
    ws2.cell(row=row, column=1, value=a["widget_name"])
    ws2.cell(row=row, column=1).font = Font(name="DM Sans", bold=True)
    ws2.cell(row=row, column=2, value=a["source_name"])
    ws2.cell(row=row, column=3, value=a["asset_class"])
    ws2.cell(row=row, column=4, value=a["start_date"])
    for c in range(1, 5):
        ws2.cell(row=row, column=c).border = thin_border
        ws2.cell(row=row, column=c).alignment = Alignment(horizontal="center")
    ws2.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws2.cell(row=row, column=2).alignment = Alignment(horizontal="left")

print("Sheet 2 (Asset Mapping) done")

# ════════════════════════════════════════
# SHEET 3: Calendar Year Returns
# ════════════════════════════════════════
ws3 = wb.create_sheet("Calendar Year Returns")
ws3.cell(row=1, column=1, value="Calendar Year Returns (All Assets)").font = title_font

# Compute yearly returns for each asset
# Return = product of (1 + monthly return) - 1 for each calendar year
yearly_data = {}  # asset_idx -> { year: return }
all_years = set()

for a_idx, a in enumerate(assets):
    yearly_data[a_idx] = {}
    vals = a["values"]
    s_start = a["series_start"]

    for d_idx in range(1, num_dates):
        data_i = d_idx - s_start
        data_prev = data_i - 1
        if (data_i >= 0 and data_i < len(vals) and
            data_prev >= 0 and data_prev < len(vals) and
            vals[data_prev] is not None and vals[data_i] is not None and
            vals[data_prev] != 0):

            monthly_ret = vals[data_i] / vals[data_prev] - 1
            yr = int(dates[d_idx][:4])
            all_years.add(yr)

            if yr not in yearly_data[a_idx]:
                yearly_data[a_idx][yr] = 1.0
            yearly_data[a_idx][yr] *= (1 + monthly_ret)

    # Convert cumulative products to returns
    for yr in yearly_data[a_idx]:
        yearly_data[a_idx][yr] -= 1.0

years_sorted = sorted(all_years)

# Headers: Year + each asset's widget name
YR_HEADER_ROW = 3
ws3.cell(row=YR_HEADER_ROW, column=1, value="Year")
for i, a in enumerate(assets):
    ws3.cell(row=YR_HEADER_ROW, column=i + 2, value=a["widget_name"])
style_header(ws3, YR_HEADER_ROW, len(assets) + 1)

ws3.column_dimensions["A"].width = 8
for i in range(len(assets)):
    col_letter = openpyxl.utils.get_column_letter(i + 2)
    ws3.column_dimensions[col_letter].width = 14

# Data rows
for y_idx, yr in enumerate(years_sorted):
    row = YR_HEADER_ROW + 1 + y_idx
    ws3.cell(row=row, column=1, value=yr)
    ws3.cell(row=row, column=1).font = Font(name="DM Sans", bold=True)
    ws3.cell(row=row, column=1).border = thin_border
    ws3.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws3.cell(row=row, column=1).number_format = '0'

    for a_idx in range(len(assets)):
        cell = ws3.cell(row=row, column=a_idx + 2)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

        if yr in yearly_data[a_idx]:
            cell.value = yearly_data[a_idx][yr]
            cell.number_format = pct_fmt

print("Sheet 3 (Calendar Year Returns) done")

# ── Save ──
out_path = "All_Widget_Data_Export.xlsx"
wb.save(out_path)
print(f"\nSaved: {out_path}")
print(f"Assets: {len(assets)}")
print(f"Date range: {dates[0]} to {dates[-1]} ({num_dates} dates)")
print(f"Years: {years_sorted[0]} to {years_sorted[-1]} ({len(years_sorted)} years)")
