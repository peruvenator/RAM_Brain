"""
QC Workbook: replicate the LIVE production widget exactly.

Data source: live_visualizer-widget.js
  (downloaded from returnstacked.com/wp-content/uploads/visualizer-widget/).
Managed Futures sleeve = PivotalPath Managed Futures Index ("MFT").

Portfolio (widget defaults):
  Core:  60% Stocks (MSCI USA) / 40% Bonds (US Agg)
  Stack: 20% total, split 25/25/25/25 -> 5% each
         Trend / Futures Yield / Gold / Merger Arb
  Fees (bp/yr):       Trend 0,  FY 0,  Gold 40, MergerArb 0
  Financing (bp/yr):  Trend 50, FY 50, Gold 50, MergerArb 50

Every monthly intermediate and every summary statistic is written as a LIVE
Excel formula so the math can be audited cell-by-cell. Change any blue input
cell on the Inputs sheet and the whole workbook recomputes.

Replicates the widget JS verbatim:
  compute():       rFixed, per-asset net/financed returns, rStacked, growth
  computeStats():  annReturn, annVol (sample n-1), maxDD, Sharpe (T-bill rf)
  computeTrackingError(): sample stdev of (stacked - fixed) * sqrt(12)
"""

import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = "live_visualizer-widget.js"

# ── Extract the embedded DATA array ──
txt = open(SRC, encoding="utf-8").read()
rows = re.findall(
    r'\{ date: "([^"]+)", stocks: ([\d.]+), bonds: ([\d.]+), gold: ([\d.]+), '
    r'trend: ([\d.]+), futuresYield: ([\d.]+), mergerArb: ([\d.]+), tbills: ([\d.]+) \}',
    txt,
)
if not rows:
    raise SystemExit("No DATA rows parsed from " + SRC)

def norm_date(d):
    m, day, y = d.split("/")
    y = int(y); y = 2000 + y if y < 50 else 1900 + y
    return f"{y:04d}-{int(m):02d}-{int(day):02d}"

DATA = [
    dict(date=norm_date(r[0]), stocks=float(r[1]), bonds=float(r[2]), gold=float(r[3]),
         trend=float(r[4]), fy=float(r[5]), marb=float(r[6]), tbills=float(r[7]))
    for r in rows
]
N = len(DATA)              # rows incl. the 12/31/99 baseline
nret = N - 1               # number of monthly returns
print(f"Parsed {N} rows ({DATA[0]['date']} -> {DATA[-1]['date']}), {nret} monthly returns")

# ── Styles ──
TEAL = "14CFA6"; NAVY = "323A46"; BLUE = "3A6A9C"; LIGHT = "EAF6F2"; GREY = "F0F1F1"
hdr_font = Font(name="DM Sans", bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor=NAVY)
inp_font = Font(name="DM Sans", bold=True, size=11, color="1F3D7A")
inp_fill = PatternFill("solid", fgColor="DCE6F5")
lbl_font = Font(name="DM Sans", bold=True, size=11, color=NAVY)
title_font = Font(name="DM Sans", bold=True, size=14, color=NAVY)
note_font = Font(name="DM Sans", italic=True, size=9, color="625C6D")
thin = Side(style="thin", color="D8DCE0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center")
right = Alignment(horizontal="right")

PCT = "0.00%"; PCT4 = "0.0000%"; NUM2 = "0.00"; LVL = "0.0000"

wb = openpyxl.Workbook()

# ───────────────────────────────────────────────────────── Inputs sheet ──
ws_in = wb.active
ws_in.title = "Inputs"
ws_in["A1"] = "Return Stacking QC - Inputs (edit blue cells)"
ws_in["A1"].font = title_font
ws_in.merge_cells("A1:C1")

inputs = [
    ("Core - Stocks weight",      "stockPct",      0.60, PCT),
    ("Core - Bonds weight",       "bondPct",       0.40, PCT),
    ("Stack size (total notional)", "stackSize",   0.20, PCT),
    ("Blend - Trend",             "trendW",        0.25, PCT),
    ("Blend - Futures Yield",     "fyW",           0.25, PCT),
    ("Blend - Gold",              "goldW",         0.25, PCT),
    ("Blend - Merger Arb",        "marbW",         0.25, PCT),
    ("Fee bp/yr - Trend",         "trendFee",      0,    "0"),
    ("Fee bp/yr - Futures Yield", "fyFee",         0,    "0"),
    ("Fee bp/yr - Gold",          "goldFee",       40,   "0"),
    ("Fee bp/yr - Merger Arb",    "marbFee",       0,    "0"),
    ("Financing bp/yr - Trend",   "trendFin",      50,   "0"),
    ("Financing bp/yr - Futures Yield", "fyFin",   50,   "0"),
    ("Financing bp/yr - Gold",    "goldFin",       50,   "0"),
    ("Financing bp/yr - Merger Arb", "marbFin",    50,   "0"),
]
ref = {}            # name -> absolute cell ref on Inputs sheet
r0 = 3
ws_in.cell(r0 - 1, 1, "Parameter").font = hdr_font
ws_in.cell(r0 - 1, 1).fill = hdr_fill
ws_in.cell(r0 - 1, 2, "Value").font = hdr_font
ws_in.cell(r0 - 1, 2).fill = hdr_fill
for i, (label, name, val, fmt) in enumerate(inputs):
    r = r0 + i
    c1 = ws_in.cell(r, 1, label); c1.font = lbl_font
    c2 = ws_in.cell(r, 2, val)
    c2.font = inp_font; c2.fill = inp_fill; c2.border = border; c2.number_format = fmt
    ref[name] = f"Inputs!$B${r}"
ws_in.column_dimensions["A"].width = 30
ws_in.column_dimensions["B"].width = 14
note_r = r0 + len(inputs) + 1
ws_in.cell(note_r, 1,
           "Defaults match the live production widget panel (gold 40bp fee, all 50bp financing).").font = note_font
ws_in.cell(note_r + 1, 1,
           "At these defaults the Stacked column reproduces the live widget output exactly.").font = note_font

# ──────────────────────────────────────────────────── Calculations sheet ──
ws = wb.create_sheet("Calculations")

# Column layout
cols = [
    ("Date", 12), ("Stocks lvl", 11), ("Bonds lvl", 11), ("Gold lvl", 11),
    ("Trend (MFT)", 12), ("FY lvl", 11), ("MergerArb lvl", 12), ("T-Bill lvl", 11),
    ("rStock", 10), ("rBond", 10), ("rTrend", 10), ("rFY", 10), ("rGold", 10),
    ("rMarb", 10), ("rTbill", 10),
    ("rFixed", 10), ("Trend net", 10), ("FY net", 10), ("Gold net", 10),
    ("Marb net", 10), ("Trend fin'd", 11), ("Alt blend", 11), ("rStacked", 11),
    ("Fixed $", 11), ("Stacked $", 11),
    ("Fix peak", 10), ("Fix DD", 10), ("Stk peak", 10), ("Stk DD", 10),
    ("Stk-Fix", 10),
]
# map header -> column letter
L = {}
for i, (h, w) in enumerate(cols):
    cidx = i + 1
    letter = get_column_letter(cidx)
    L[h] = letter
    cell = ws.cell(1, cidx, h)
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
    ws.column_dimensions[letter].width = w

# convenience
S, B, G, T, F, M, TB = (L["Stocks lvl"], L["Bonds lvl"], L["Gold lvl"], L["Trend (MFT)"],
                        L["FY lvl"], L["MergerArb lvl"], L["T-Bill lvl"])
rS, rB, rT, rF, rG, rM, rTb = (L["rStock"], L["rBond"], L["rTrend"], L["rFY"], L["rGold"],
                               L["rMarb"], L["rTbill"])
rFix = L["rFixed"]; tNet = L["Trend net"]; fNet = L["FY net"]; gNet = L["Gold net"]
mNet = L["Marb net"]; tFin = L["Trend fin'd"]; blend = L["Alt blend"]; rStk = L["rStacked"]
fixD = L["Fixed $"]; stkD = L["Stacked $"]
fpk = L["Fix peak"]; fdd = L["Fix DD"]; spk = L["Stk peak"]; sdd = L["Stk DD"]; diff = L["Stk-Fix"]

# data + formula rows.  Excel row 2 = baseline (12/31/99), rows 3..N+1 = returns
def er(i):  # excel row for DATA index i
    return i + 2

for i, d in enumerate(DATA):
    r = er(i)
    ws.cell(r, 1, d["date"]).alignment = center
    ws.cell(r, 2, d["stocks"]).number_format = LVL
    ws.cell(r, 3, d["bonds"]).number_format = LVL
    ws.cell(r, 4, d["gold"]).number_format = LVL
    ws.cell(r, 5, d["trend"]).number_format = LVL
    ws.cell(r, 6, d["fy"]).number_format = LVL
    ws.cell(r, 7, d["marb"]).number_format = LVL
    ws.cell(r, 8, d["tbills"]).number_format = LVL

    if i == 0:
        # baseline growth = $100, no returns on the first row
        ws[f"{fixD}{r}"] = 100
        ws[f"{stkD}{r}"] = 100
        ws[f"{fpk}{r}"] = f"={fixD}{r}"
        ws[f"{spk}{r}"] = f"={stkD}{r}"
        ws[f"{fdd}{r}"] = 0
        ws[f"{sdd}{r}"] = 0
        for col in (fixD, stkD):
            ws[f"{col}{r}"].number_format = "0.0000"
        for col in (fdd, sdd):
            ws[f"{col}{r}"].number_format = PCT
        continue

    p = r - 1  # previous excel row
    # monthly returns
    ws[f"{rS}{r}"]  = f"={S}{r}/{S}{p}-1"
    ws[f"{rB}{r}"]  = f"={B}{r}/{B}{p}-1"
    ws[f"{rT}{r}"]  = f"={T}{r}/{T}{p}-1"
    ws[f"{rF}{r}"]  = f"={F}{r}/{F}{p}-1"
    ws[f"{rG}{r}"]  = f"={G}{r}/{G}{p}-1"
    ws[f"{rM}{r}"]  = f"={M}{r}/{M}{p}-1"
    ws[f"{rTb}{r}"] = f"={TB}{r}/{TB}{p}-1"
    # fixed mix
    ws[f"{rFix}{r}"] = f"={ref['stockPct']}*{rS}{r}+{ref['bondPct']}*{rB}{r}"
    # per-asset net / financed (verbatim widget formulas; bp/10000/12 monthly)
    ws[f"{tNet}{r}"] = f"={rT}{r}-{ref['trendFee']}/10000/12"
    ws[f"{fNet}{r}"] = f"={rF}{r}-{ref['fyFee']}/10000/12-({rTb}{r}+{ref['fyFin']}/10000/12)"
    ws[f"{gNet}{r}"] = f"={rG}{r}-{ref['goldFee']}/10000/12-({rTb}{r}+{ref['goldFin']}/10000/12)"
    ws[f"{mNet}{r}"] = f"={rM}{r}-{ref['marbFee']}/10000/12-({rTb}{r}+{ref['marbFin']}/10000/12)"
    ws[f"{tFin}{r}"] = f"={tNet}{r}-({rTb}{r}+{ref['trendFin']}/10000/12)"
    # alt blend & stacked
    ws[f"{blend}{r}"] = (f"={ref['trendW']}*{tFin}{r}+{ref['fyW']}*{fNet}{r}"
                          f"+{ref['goldW']}*{gNet}{r}+{ref['marbW']}*{mNet}{r}")
    ws[f"{rStk}{r}"]  = f"={rFix}{r}+{ref['stackSize']}*{blend}{r}"
    # growth of $100
    ws[f"{fixD}{r}"] = f"={fixD}{p}*(1+{rFix}{r})"
    ws[f"{stkD}{r}"] = f"={stkD}{p}*(1+{rStk}{r})"
    # drawdown helpers
    ws[f"{fpk}{r}"] = f"=MAX({fpk}{p},{fixD}{r})"
    ws[f"{fdd}{r}"] = f"=({fpk}{r}-{fixD}{r})/{fpk}{r}"
    ws[f"{spk}{r}"] = f"=MAX({spk}{p},{stkD}{r})"
    ws[f"{sdd}{r}"] = f"=({spk}{r}-{stkD}{r})/{spk}{r}"
    ws[f"{diff}{r}"] = f"={rStk}{r}-{rFix}{r}"

    # number formats
    for col in (rS, rB, rT, rF, rG, rM, rTb, rFix, tNet, fNet, gNet, mNet, tFin, blend, rStk, diff):
        ws[f"{col}{r}"].number_format = PCT4
    for col in (fixD, stkD, fpk, spk):
        ws[f"{col}{r}"].number_format = "0.0000"
    for col in (fdd, sdd):
        ws[f"{col}{r}"].number_format = PCT

ws.freeze_panes = "B2"

# ranges used by Results
first, last = er(1), er(N - 1)
rng_rFix  = f"Calculations!${rFix}${first}:${rFix}${last}"
rng_rStk  = f"Calculations!${rStk}${first}:${rStk}${last}"
rng_fdd   = f"Calculations!${fdd}${first}:${fdd}${last}"
rng_sdd   = f"Calculations!${sdd}${first}:${sdd}${last}"
rng_diff  = f"Calculations!${diff}${first}:${diff}${last}"
fixD_first = f"Calculations!${fixD}${er(0)}"; fixD_last = f"Calculations!${fixD}${last}"
stkD_first = f"Calculations!${stkD}${er(0)}"; stkD_last = f"Calculations!${stkD}${last}"
tb_first = f"Calculations!${TB}${er(0)}"; tb_last = f"Calculations!${TB}${last}"

# ──────────────────────────────────────────────────────── Results sheet ──
wr = wb.create_sheet("Results", 0)  # make it the first tab
wr["A1"] = "Return Stacking - QC Results"
wr["A1"].font = title_font
wr.merge_cells("A1:D1")
wr["A2"] = f"Window: {DATA[1]['date']} to {DATA[-1]['date']}  ({nret} monthly returns)"
wr["A2"].font = note_font

hdr = ["Metric", "Stock/Bond", "Stacked", "Difference"]
for j, h in enumerate(hdr):
    c = wr.cell(4, j + 1, h); c.font = hdr_font; c.fill = hdr_fill
    c.alignment = center if j else Alignment(horizontal="left"); c.border = border

NEXP = f"(12/{nret})"
# Annualized Return
wr.cell(5, 1, "Annualized Return").font = lbl_font
wr["B5"] = f"=({fixD_last}/{fixD_first})^{NEXP}-1"
wr["C5"] = f"=({stkD_last}/{stkD_first})^{NEXP}-1"
wr["D5"] = "=C5-B5"
# Annualized Volatility (sample stdev * sqrt(12))
wr.cell(6, 1, "Annualized Volatility").font = lbl_font
wr["B6"] = f"=STDEV.S({rng_rFix})*SQRT(12)"
wr["C6"] = f"=STDEV.S({rng_rStk})*SQRT(12)"
wr["D6"] = "=C6-B6"
# Maximum Drawdown (most negative -> report as negative)
wr.cell(7, 1, "Maximum Drawdown").font = lbl_font
wr["B7"] = f"=-MAX({rng_fdd})"
wr["C7"] = f"=-MAX({rng_sdd})"
wr["D7"] = "=C7-B7"
# Risk-free (T-bill annualized) - shown for transparency
wr.cell(8, 1, "Risk-free (T-Bill ann.)").font = lbl_font
wr["B8"] = f"=({tb_last}/{tb_first})^{NEXP}-1"
wr["C8"] = "=B8"
# Sharpe Ratio
wr.cell(9, 1, "Sharpe Ratio").font = lbl_font
wr["B9"] = "=(B5-B8)/B6"
wr["C9"] = "=(C5-C8)/C6"
wr["D9"] = "=C9-B9"
# Tracking Error
wr.cell(10, 1, "Tracking Error").font = lbl_font
wr["C10"] = f"=STDEV.S({rng_diff})*SQRT(12)"

for r in range(5, 11):
    for col in ("B", "C", "D"):
        cell = wr[f"{col}{r}"]
        cell.alignment = right; cell.border = border
        cell.number_format = NUM2 if r == 9 else PCT
wr.column_dimensions["A"].width = 24
for col in ("B", "C", "D"):
    wr.column_dimensions[col].width = 13

wr.cell(12, 1, "Live production widget output (returnstacked.com) for cross-check:").font = lbl_font
xcheck = [
    ("Annualized Return", "6.75%", "8.09%", "+1.34%"),
    ("Annualized Volatility", "9.46%", "9.66%", "+0.20%"),
    ("Maximum Drawdown", "-32.54%", "-30.84%", "+1.70%"),
    ("Sharpe Ratio", "0.51", "0.64", "+0.13"),
]
for i, row in enumerate(xcheck):
    for j, v in enumerate(row):
        c = wr.cell(13 + i, j + 1, v)
        c.font = note_font
        if j: c.alignment = right

wr.cell(18, 1,
        "Formulas mirror the widget's compute()/computeStats() line-for-line. "
        "Vol & TE use sample stdev (n-1); Sharpe rf = T-Bill CAGR over window.").font = note_font

OUT = "QC_LiveWidget_Verification.xlsx"
wb.save(OUT)
print("Saved", OUT)
