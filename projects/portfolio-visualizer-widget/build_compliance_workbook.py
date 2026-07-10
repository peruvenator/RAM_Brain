"""Build Compliance_60_40_Stacked.xlsx.

Mirrors the widget's exact computation methodology in Excel cells so each
derived value is fully traceable back to the underlying index series.

Sheets:
  - Index Levels:  raw underlying price levels + formula-driven 60/40 level,
                   60/40 + 30% stack - 30% T-Bills level, and difference.
                   Helper columns for running peak / drawdown.
  - Monthly Returns: every cell is a formula referencing the level series
                     (underlyings) and the same-row underlyings (derived).
  - Statistics: five-metric table, all cells are formulas. Methodology
                matches the widget's `computeStats()` exactly:
                  - annualized return = (cum_return)^(12/n) - 1
                  - annualized vol    = STDEV.S(returns) * SQRT(12)
                  - max drawdown      = MAX((peak - level) / peak)
                  - sharpe            = (annReturn - rfAnn) / vol
                                        with rfAnn = AVERAGE(tbill_returns) * 12
                                        (arithmetic mean, per widget code)
                  - tracking error    = STDEV.S(stacked - benchmark) * SQRT(12)
"""
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

OUT = "Compliance_60_40_Stacked.xlsx"
BASE_DATE = "1999-12-31"

with open("data/indices_compact.json") as f:
    d = json.load(f)
dates = d["dates"]
series = d["series"]
base_idx = dates.index(BASE_DATE)

def slice_series(key):
    s = series[key]
    full = [None] * s["start"] + s["values"]
    return [full[base_idx + i] for i in range(len(dates) - base_idx)]

eq = slice_series("MSCI USA")
agg = slice_series("Bloomberg US Aggregate Bond")
sht = slice_series("Treasury Bill")
mf = slice_series("MFT")
ma = slice_series("EVDMER")
gold = slice_series("Gold (Spot)")
dates_out = dates[base_idx:]
n = len(dates_out)
last_row = 2 + n          # data starts row 3, ends row last_row
n_returns = n - 1
returns_last_row = 2 + n_returns

wb = openpyxl.Workbook()

# ───────── Sheet 1: Index Levels ─────────
ws = wb.active
ws.title = "Index Levels"

headers1 = [
    "Date",
    "U.S. Large Cap Equities",
    "U.S. Aggregate Bonds",
    "Short Treasury",
    "Managed Futures CTA",
    "Merger Arbitrage",
    "Gold",
    "60/40 Portfolio (Level)",
    "60/40 + 30% Stack - 30% T-Bills (Level)",
    "Difference: Stacked - 60/40 (Level)",
]
headers2 = [
    "Underlying index",
    "MSCI USA (NDDUUS) Total Return",
    "Bloomberg US Aggregate Bond (LBUSTRUU)",
    "Bloomberg US T-Bill Total Return (LD12TRUU)",
    "PivotalPath Managed Futures Index (MFT)",
    "PivotalPath Event Driven: Merger Arbitrage (EVDMER)",
    "Gold Spot USD (XAU Curncy)",
    "60% Equity + 40% Aggregate Bonds",
    "60% Equity + 40% Bonds + 10% MF + 10% MA + 10% Gold - 30% T-Bills",
    "Column I minus Column H",
]
helper_headers = [
    "Peak (60/40)",   # K
    "DD (60/40)",     # L
    "Peak (Stacked)", # M
    "DD (Stacked)",   # N
]

bold = Font(bold=True)
header_fill = PatternFill("solid", fgColor="14CFA6")  # brand teal
for c, h in enumerate(headers1 + helper_headers, 1):
    cell = ws.cell(1, c, h)
    cell.font = Font(bold=True, color="FFFFFF" if c <= 10 else "000000")
    if c <= 10:
        cell.fill = header_fill
for c, h in enumerate(headers2, 1):
    ws.cell(2, c, h).font = Font(italic=True)

# Row 3: base date with raw underlying values; H/I = 100; J = 0
ws.cell(3, 1, dates_out[0])
ws.cell(3, 2, eq[0])
ws.cell(3, 3, agg[0])
ws.cell(3, 4, sht[0])
ws.cell(3, 5, mf[0])
ws.cell(3, 6, ma[0])
ws.cell(3, 7, gold[0])
ws.cell(3, 8, 100)                            # H
ws.cell(3, 9, 100)                            # I
ws.cell(3, 10, "=I3-H3")                      # J
ws.cell(3, 11, "=H3")                         # K peak
ws.cell(3, 12, "=(K3-H3)/K3")                 # L dd
ws.cell(3, 13, "=I3")                         # M peak
ws.cell(3, 14, "=(M3-I3)/M3")                 # N dd

underlyings = [eq, agg, sht, mf, ma, gold]
for i in range(1, n):
    r = 3 + i
    ws.cell(r, 1, dates_out[i])
    for c_off, vals in enumerate(underlyings):
        ws.cell(r, 2 + c_off, vals[i])
    # H: 60/40 level. r_t = 0.6*(B_r/B_{r-1}-1) + 0.4*(C_r/C_{r-1}-1)
    ws.cell(r, 8, f"=H{r-1}*(1+0.6*(B{r}/B{r-1}-1)+0.4*(C{r}/C{r-1}-1))")
    # I: 60/40 + 10% MF + 10% MA + 10% Gold - 30% T-Bills
    ws.cell(
        r, 9,
        f"=I{r-1}*(1"
        f"+0.6*(B{r}/B{r-1}-1)"
        f"+0.4*(C{r}/C{r-1}-1)"
        f"+0.1*(E{r}/E{r-1}-1)"
        f"+0.1*(F{r}/F{r-1}-1)"
        f"+0.1*(G{r}/G{r-1}-1)"
        f"-0.3*(D{r}/D{r-1}-1))"
    )
    # J: level difference
    ws.cell(r, 10, f"=I{r}-H{r}")
    # Helpers for max drawdown
    ws.cell(r, 11, f"=MAX($H$3:H{r})")
    ws.cell(r, 12, f"=(K{r}-H{r})/K{r}")
    ws.cell(r, 13, f"=MAX($I$3:I{r})")
    ws.cell(r, 14, f"=(M{r}-I{r})/M{r}")

# Column widths
widths = [12, 22, 22, 20, 24, 24, 16, 26, 42, 32, 16, 14, 16, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "B3"

# Number formats
for r in range(3, last_row + 1):
    for c in range(2, 11):
        ws.cell(r, c).number_format = "#,##0.0000"
    for c in (12, 14):
        ws.cell(r, c).number_format = "0.00%"

# ───────── Sheet 2: Monthly Returns ─────────
ws2 = wb.create_sheet("Monthly Returns")

ret_headers1 = [
    "Date",
    "U.S. Large Cap Equities",
    "U.S. Aggregate Bonds",
    "Short Treasury",
    "Managed Futures CTA",
    "Merger Arbitrage",
    "Gold",
    "60/40 Portfolio (Monthly Return)",
    "60/40 + 30% Stack - 30% T-Bills (Monthly Return)",
    "Active Return: Stacked - 60/40",
]
ret_headers2 = list(headers2)
ret_headers2[7] = "0.6*r(Eq) + 0.4*r(Agg)"
ret_headers2[8] = "0.6*r(Eq) + 0.4*r(Agg) + 0.1*r(MF) + 0.1*r(MA) + 0.1*r(Gold) - 0.3*r(T-Bill)"
ret_headers2[9] = "Column I minus Column H"

for c, h in enumerate(ret_headers1, 1):
    cell = ws2.cell(1, c, h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
for c, h in enumerate(ret_headers2, 1):
    ws2.cell(2, c, h).font = Font(italic=True)

for i in range(1, n):
    r = 2 + i               # row 3 corresponds to first return (i=1)
    src_r_curr = 3 + i      # Index Levels row for current month
    src_r_prev = 2 + i      # Index Levels row for prior month
    ws2.cell(r, 1, dates_out[i])
    for c_off, col_letter in enumerate(["B", "C", "D", "E", "F", "G"]):
        ws2.cell(r, 2 + c_off,
                 f"='Index Levels'!{col_letter}{src_r_curr}/'Index Levels'!{col_letter}{src_r_prev}-1")
    # 60/40 monthly return: 0.6 * eq_ret + 0.4 * agg_ret  (referencing same-row underlyings)
    ws2.cell(r, 8, f"=0.6*B{r}+0.4*C{r}")
    # Stacked monthly return
    ws2.cell(r, 9, f"=0.6*B{r}+0.4*C{r}+0.1*E{r}+0.1*F{r}+0.1*G{r}-0.3*D{r}")
    # Active return: stacked minus 60/40  (=> simplifies to 0.1*E + 0.1*F + 0.1*G - 0.3*D)
    ws2.cell(r, 10, f"=I{r}-H{r}")

widths2 = [12, 22, 22, 20, 24, 24, 16, 30, 46, 28]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "B3"
for r in range(3, returns_last_row + 1):
    for c in range(2, 11):
        ws2.cell(r, c).number_format = "0.0000%"

# ───────── Sheet 3: Statistics ─────────
ws3 = wb.create_sheet("Statistics")
ws3["A1"] = "Summary Statistics (formulas mirror the widget's computeStats methodology)"
ws3["A1"].font = Font(bold=True, size=12)

# Column headers
ws3.cell(3, 1, "Metric").font = bold
ws3.cell(3, 2, "60/40 Portfolio").font = Font(bold=True, color="FFFFFF")
ws3.cell(3, 2).fill = header_fill
ws3.cell(3, 3, "60/40 + 30% Stack - 30% T-Bills").font = Font(bold=True, color="FFFFFF")
ws3.cell(3, 3).fill = header_fill
ws3.cell(3, 4, "Difference (Stacked - 60/40)").font = Font(bold=True, color="FFFFFF")
ws3.cell(3, 4).fill = header_fill

# References
H_last = f"'Index Levels'!H{last_row}"
I_last = f"'Index Levels'!I{last_row}"
H_first = "'Index Levels'!H3"
I_first = "'Index Levels'!I3"
H_returns = f"'Monthly Returns'!H3:H{returns_last_row}"
I_returns = f"'Monthly Returns'!I3:I{returns_last_row}"
J_returns = f"'Monthly Returns'!J3:J{returns_last_row}"
D_returns = f"'Monthly Returns'!D3:D{returns_last_row}"   # T-Bill monthly returns
DD_H = f"'Index Levels'!L3:L{last_row}"
DD_I = f"'Index Levels'!N3:N{last_row}"
N_R = n_returns

# Row 4: Annualized Return
ws3.cell(4, 1, "Annualized Return")
ws3.cell(4, 2, f"=({H_last}/{H_first})^(12/{N_R})-1")
ws3.cell(4, 3, f"=({I_last}/{I_first})^(12/{N_R})-1")
ws3.cell(4, 4, "=C4-B4")

# Row 5: Annualized Volatility
ws3.cell(5, 1, "Annualized Volatility")
ws3.cell(5, 2, f"=STDEV.S({H_returns})*SQRT(12)")
ws3.cell(5, 3, f"=STDEV.S({I_returns})*SQRT(12)")
ws3.cell(5, 4, "=C5-B5")

# Row 6: Max Drawdown
ws3.cell(6, 1, "Maximum Drawdown")
ws3.cell(6, 2, f"=MAX({DD_H})")
ws3.cell(6, 3, f"=MAX({DD_I})")
ws3.cell(6, 4, "=C6-B6")

# Row 7: Sharpe Ratio  -- (annReturn - rfAnn) / vol, rfAnn = AVERAGE(tbill_returns)*12
ws3.cell(7, 1, "Sharpe Ratio")
ws3.cell(7, 2, f"=(B4-AVERAGE({D_returns})*12)/B5")
ws3.cell(7, 3, f"=(C4-AVERAGE({D_returns})*12)/C5")
ws3.cell(7, 4, "=C7-B7")

# Row 8: Annualized Tracking Error (vs 60/40)
ws3.cell(8, 1, "Annualized Tracking Error (vs 60/40)")
ws3.cell(8, 2, 0)
ws3.cell(8, 3, f"=STDEV.S({J_returns})*SQRT(12)")
ws3.cell(8, 4, "=C8-B8")

# Formats
pct_rows = (4, 5, 6, 8)
ratio_rows = (7,)
for rr in pct_rows:
    for c in (2, 3, 4):
        ws3.cell(rr, c).number_format = "0.00%"
for rr in ratio_rows:
    for c in (2, 3, 4):
        ws3.cell(rr, c).number_format = "0.000"

ws3.column_dimensions["A"].width = 42
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 34
ws3.column_dimensions["D"].width = 28

# Methodology block
ws3.cell(10, 1, "Methodology (mirrors widget computeStats):").font = bold
methodology = [
    "  Annualized Return  = (End Level / Start Level)^(12 / n_months) - 1",
    "                       where n_months = number of monthly returns in the sample.",
    "  Annualized Volatility = STDEV.S(monthly returns) * SQRT(12)  [sample standard deviation, n-1].",
    "  Maximum Drawdown   = MAX over t of (Peak(0..t) - Level(t)) / Peak(0..t).",
    "                       Peak helper columns are on 'Index Levels' (K and M); drawdown helpers L and N.",
    "  Sharpe Ratio       = (Annualized Return - Risk-Free Annual) / Annualized Volatility,",
    "                       Risk-Free Annual = AVERAGE(monthly T-Bill returns) * 12  (arithmetic, per widget).",
    "  Annualized Tracking Error = STDEV.S(stacked monthly returns - 60/40 monthly returns) * SQRT(12),",
    "                       equivalent to STDEV.S of column J on 'Monthly Returns' times SQRT(12).",
]
for i, line in enumerate(methodology, 11):
    ws3.cell(i, 1, line)

# ───────── Sheet 4: Notes ─────────
ws4 = wb.create_sheet("Notes")
notes = [
    "Workbook purpose: provide a fully transparent, compliance-ready record of",
    "the independent index price levels used, the portfolio constructions, and the",
    "summary statistics, with every derived cell expressed as an Excel formula.",
    "",
    "Underlying series (Sheet 'Index Levels', columns B-G):",
    "  B  U.S. Large Cap Equities      MSCI USA (NDDUUS) Total Return Index",
    "  C  U.S. Aggregate Bonds         Bloomberg US Aggregate Bond TR (LBUSTRUU)",
    "  D  Short Treasury               Bloomberg US Treasury Bill TR (LD12TRUU)",
    "  E  Managed Futures CTA          PivotalPath Managed Futures Index (MFT)",
    "  F  Merger Arbitrage             PivotalPath Event Driven: Merger Arbitrage Index (EVDMER)",
    "  G  Gold                         Gold Spot USD (XAU Curncy)",
    "",
    "Values shown for columns B-G are the actual underlying total-return index",
    "levels as reported by the respective providers. No rebasing has been applied",
    "to the underlying series.",
    "",
    "Derived portfolios (Sheet 'Index Levels', columns H-J), monthly rebalanced,",
    "growth-of-100 starting at the base date 1999-12-31:",
    "  H  60/40                       0.6 * r(Equity) + 0.4 * r(Aggregate Bonds)",
    "  I  60/40 + 30% Stack           0.6 * r(Equity) + 0.4 * r(Agg) + 0.1 * r(MF) + 0.1 * r(MA) + 0.1 * r(Gold) - 0.3 * r(T-Bill)",
    "  J  Difference (Level)          Column I minus Column H",
    "",
    "Each row in columns H, I, J is an Excel formula referencing the prior row's",
    "derived level and the current/prior underlying levels for the relevant series,",
    "so the construction is fully traceable.",
    "",
    "Monthly return sheet (Sheet 'Monthly Returns'):",
    "  - Columns B-G are the monthly returns of the underlying series, computed as",
    "    'Index Levels'!X_t / 'Index Levels'!X_{t-1} - 1.",
    "  - Columns H, I, J use the formulas shown in the header subtitles.",
    "",
    "Statistics (Sheet 'Statistics'): formulas mirror the visualizer widget's",
    "computeStats() function. See the methodology block on that sheet.",
    "",
    "Note on U.S. Large-Cap proxy: the widget uses MSCI USA Total Return (NDDUUS)",
    "as its 'U.S. Large Cap Equities' series, which is the proxy reproduced here.",
    "If S&P 500 Total Return (SPX) is required for compliance, provide the monthly",
    "SPX TR levels and the workbook can be regenerated with that series substituted.",
    "",
    f"Data span: {dates_out[0]} to {dates_out[-1]} ({n} monthly observations, {n_returns} monthly returns).",
    "Base date for derived growth-of-100 series: " + BASE_DATE + ".",
]
for i, line in enumerate(notes, 1):
    ws4.cell(i, 1, line)
ws4.column_dimensions["A"].width = 110

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  Index Levels rows: 3..{last_row} ({n} months)")
print(f"  Monthly Returns rows: 3..{returns_last_row} ({n_returns} months)")
print(f"  Stats: 5 metrics x 3 columns, all formulas")
