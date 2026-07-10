"""Build Compliance_Simple_Visualizer_Match.xlsx.

Reproduces the 'Simple Visualiser Old with carry.html' widget output exactly,
in Excel, with every derived cell expressed as a formula referencing:
  - The same underlying index price levels embedded in that widget.
  - Live input cells for weights, fees, and financing spreads (basis points).

The stacked return formula is the widget's compute() formula verbatim:

  rFixed        = sw * rStock + bw * rBond
  trendNet      = rTrend - trendFee/10000/12
  trendFinanced = trendNet - (rTbill + trendFinancing/10000/12)
  fyNet         = rFy   - fyFee/10000/12   - (rTbill + fyFinancing/10000/12)
  goldNet       = rGold - goldFee/10000/12 - (rTbill + goldFinancing/10000/12)
  marbNet       = rMarb - marbFee/10000/12 - (rTbill + marbFinancing/10000/12)
  rStacked      = rFixed + ss * (tw*trendFinanced + fw*fyNet + gw*goldNet + mw*marbNet)

Default Inputs sheet values reflect the user's stated weight scheme:
  60% Stocks, 40% Bonds, 30% Stack Size,
  Stack composition: 33.34% Managed Futures, 33.33% Merger Arb, 33.33% Gold
  Fees/financing default to the widget's defaults (40 bp Gold fee, 50 bp
  financing on every overlay).  Adjust on the Inputs sheet to match any
  alternative widget run; every downstream formula recalculates.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "Compliance_Simple_Visualizer_Match.xlsx"

# Use the LIVE widget data pulled from returnstacked.com.
# (The on-disk 'Simple Visualiser Old with carry.html' is stale: it has SG
# Trend Index labeled as MFT.  The live widget JS uses actual MFT data.)
with open("live_visualizer_data.json") as f:
    DATA = json.load(f)

n = len(DATA)
n_returns = n - 1

wb = openpyxl.Workbook()

# ───────── Sheet 1: Inputs ─────────
ws = wb.active
ws.title = "Inputs"

teal = "14CFA6"
navy = "323A46"
header_fill = PatternFill("solid", fgColor=teal)
section_fill = PatternFill("solid", fgColor=navy)
input_fill = PatternFill("solid", fgColor="FFF7C2")  # pale yellow for editable

ws["A1"] = "Inputs (edit yellow cells to match any widget configuration)"
ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
ws["A1"].fill = section_fill
ws.merge_cells("A1:C1")

inputs = [
    # (label, cell ref, default value, units, comment)
    ("Weights (% of total)",      None, None, None, None),
    ("  Stocks (S&P 500 TR)",     "B3",  60,   "%",   None),
    ("  Bonds (Bloomberg US Agg)", "B4",  40,   "%",   None),
    ("  Stack Size",              "B5",  30,   "%",   "Total notional alt overlay on top of stock/bond"),
    ("Stack composition (% of stack)", None, None, None, None),
    ("  Managed Futures (Trend)", "B7",  33.34, "%",  None),
    ("  Futures Yield (Carry)",   "B8",  0,    "%",   None),
    ("  Gold",                    "B9",  33.33, "%",  None),
    ("  Merger Arbitrage",        "B10", 33.33, "%",  None),
    ("Fees (basis points)",       None, None, None, None),
    ("  Managed Futures fee",     "B12", 0,    "bp",  "PivotalPath MFT is reported net of fees"),
    ("  Futures Yield fee",       "B13", 0,    "bp",  None),
    ("  Gold fee",                "B14", 40,   "bp",  "Widget default"),
    ("  Merger Arb fee",          "B15", 0,    "bp",  "PivotalPath EVDMER is reported net of fees"),
    ("Financing spreads (basis points, in addition to T-Bills)", None, None, None, None),
    ("  Managed Futures financing", "B17", 50, "bp",  "Widget default"),
    ("  Futures Yield financing",   "B18", 50, "bp",  "Widget default"),
    ("  Gold financing",            "B19", 50, "bp",  "Widget default"),
    ("  Merger Arb financing",      "B20", 50, "bp",  "Widget default"),
]

bold = Font(bold=True)
italic = Font(italic=True, color="555555")

for i, (lbl, cell, val, units, note) in enumerate(inputs, 2):
    ws.cell(i, 1, lbl)
    if cell is None:
        ws.cell(i, 1).font = bold
        ws.cell(i, 1).fill = header_fill
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        ws.cell(i, 1).font = Font(bold=True, color="FFFFFF")
    else:
        ws.cell(i, 2, val)
        ws.cell(i, 2).fill = input_fill
        ws.cell(i, 2).number_format = "0.00"
        ws.cell(i, 3, units)
        if note:
            ws.cell(i, 4, note); ws.cell(i, 4).font = italic

# Defined names so downstream formulas read cleanly
from openpyxl.workbook.defined_name import DefinedName
name_map = {
    "sw": "$B$3", "bw": "$B$4", "ss": "$B$5",
    "tw": "$B$7", "fw": "$B$8", "gw": "$B$9", "mw": "$B$10",
    "trendFee": "$B$12", "fyFee": "$B$13", "goldFee": "$B$14", "marbFee": "$B$15",
    "trendFinancing": "$B$17", "fyFinancing": "$B$18",
    "goldFinancing": "$B$19", "marbFinancing": "$B$20",
}
for name, ref in name_map.items():
    dn = DefinedName(name, attr_text=f"Inputs!{ref}")
    wb.defined_names[name] = dn

ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 6
ws.column_dimensions["D"].width = 56

# Quick consistency check displayed as text
ws.cell(22, 1, "Sanity checks").font = bold
ws.cell(23, 1, "Stocks + Bonds (should equal 100%)")
ws.cell(23, 2, "=B3+B4")
ws.cell(23, 2).number_format = "0.00"
ws.cell(24, 1, "Stack composition (should equal 100% if Stack Size > 0)")
ws.cell(24, 2, "=B7+B8+B9+B10")
ws.cell(24, 2).number_format = "0.00"
ws.cell(25, 1, "Net long exposure: Stocks + Bonds + Stack Size = total notional")
ws.cell(25, 2, "=B3+B4+B5")
ws.cell(25, 2).number_format = "0.00"

# ───────── Sheet 2: Index Levels ─────────
ws2 = wb.create_sheet("Index Levels")
hdr_top = [
    "Date",
    "Stocks (S&P 500 TR)",
    "Bonds (Bloomberg US Agg)",
    "Gold (Spot)",
    "Managed Futures (PivotalPath MFT)",
    "Futures Yield (Carry)",
    "Merger Arb (PivotalPath EVDMER)",
    "Short Treasury T-Bills",
    "Stock/Bond Portfolio (Level)",
    "Stacked Portfolio (Level)",
    "Difference (Stacked - Stock/Bond)",
]
hdr_sub = [
    "Bloomberg tickers / providers",
    "SPX (S&P 500 Total Return)",
    "LBUSTRUU",
    "XAU Currency",
    "MFT",
    "Futures Yield (Carry)",
    "EVDMER",
    "LD12TRUUU (Bloomberg Short Treasury TR)",
    "sw * r(Stock) + bw * r(Bond)",
    "Stock/Bond + ss * Sum(weight_i * (R_alt_i - fee_i/12 - (R_tbill + fin_i/12)))",
    "Column J minus Column I",
]
for c, h in enumerate(hdr_top, 1):
    cell = ws2.cell(1, c, h)
    cell.font = Font(bold=True, color="FFFFFF"); cell.fill = header_fill
for c, h in enumerate(hdr_sub, 1):
    ws2.cell(2, c, h).font = italic

# Row 3 = base date 12/31/99
ws2.cell(3, 1, DATA[0]["date"])
ws2.cell(3, 2, DATA[0]["stocks"])
ws2.cell(3, 3, DATA[0]["bonds"])
ws2.cell(3, 4, DATA[0]["gold"])
ws2.cell(3, 5, DATA[0]["trend"])
ws2.cell(3, 6, DATA[0]["futuresYield"])
ws2.cell(3, 7, DATA[0]["mergerArb"])
ws2.cell(3, 8, DATA[0]["tbills"])
ws2.cell(3, 9, 100)
ws2.cell(3, 10, 100)
ws2.cell(3, 11, "=J3-I3")

# Helper columns for max drawdown
hdr_helper = ["Peak (S/B)", "DD (S/B)", "Peak (Stacked)", "DD (Stacked)"]
for c, h in enumerate(hdr_helper, 12):
    ws2.cell(1, c, h).font = Font(bold=True)
ws2.cell(3, 12, "=I3")
ws2.cell(3, 13, "=(L3-I3)/L3")
ws2.cell(3, 14, "=J3")
ws2.cell(3, 15, "=(N3-J3)/N3")

# Convert per-month basis-point cost to a monthly decimal: bp/10000/12
# In Excel formula: <bp_cell>/10000/12.   sw/100, bw/100 etc. for percentages.
for i in range(1, n):
    r = 3 + i
    rec = DATA[i]
    ws2.cell(r, 1, rec["date"])
    ws2.cell(r, 2, rec["stocks"])
    ws2.cell(r, 3, rec["bonds"])
    ws2.cell(r, 4, rec["gold"])
    ws2.cell(r, 5, rec["trend"])
    ws2.cell(r, 6, rec["futuresYield"])
    ws2.cell(r, 7, rec["mergerArb"])
    ws2.cell(r, 8, rec["tbills"])
    pr = r - 1

    # Stock/Bond monthly return:
    #   rFixed = (sw/100) * (B/Bp - 1) + (bw/100) * (C/Cp - 1)
    rFixed = f"((sw/100)*(B{r}/B{pr}-1)+(bw/100)*(C{r}/C{pr}-1))"
    # Underlying monthly returns
    rStk  = f"(B{r}/B{pr}-1)"
    rBnd  = f"(C{r}/C{pr}-1)"
    rGld  = f"(D{r}/D{pr}-1)"
    rTrd  = f"(E{r}/E{pr}-1)"
    rFy   = f"(F{r}/F{pr}-1)"
    rMrb  = f"(G{r}/G{pr}-1)"
    rTb   = f"(H{r}/H{pr}-1)"
    # Per-asset net / financed returns (match simple visualizer compute() exactly)
    trendFinanced = f"({rTrd}-trendFee/10000/12-({rTb}+trendFinancing/10000/12))"
    fyNet         = f"({rFy}-fyFee/10000/12-({rTb}+fyFinancing/10000/12))"
    goldNet       = f"({rGld}-goldFee/10000/12-({rTb}+goldFinancing/10000/12))"
    marbNet       = f"({rMrb}-marbFee/10000/12-({rTb}+marbFinancing/10000/12))"
    stack_excess = (
        f"(ss/100)*((tw/100)*{trendFinanced}+(fw/100)*{fyNet}+(gw/100)*{goldNet}+(mw/100)*{marbNet})"
    )
    rStacked = f"{rFixed}+{stack_excess}"

    # Level columns: prev * (1 + r)
    ws2.cell(r, 9,  f"=I{pr}*(1+{rFixed})")
    ws2.cell(r, 10, f"=J{pr}*(1+{rStacked})")
    ws2.cell(r, 11, f"=J{r}-I{r}")
    # DD helpers
    ws2.cell(r, 12, f"=MAX($I$3:I{r})")
    ws2.cell(r, 13, f"=(L{r}-I{r})/L{r}")
    ws2.cell(r, 14, f"=MAX($J$3:J{r})")
    ws2.cell(r, 15, f"=(N{r}-J{r})/N{r}")

# Widths and formats
widths2 = [12, 22, 24, 14, 30, 22, 30, 22, 26, 26, 28, 14, 12, 16, 14]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "B3"
last_row_levels = 2 + n
for r in range(3, last_row_levels + 1):
    for c in range(2, 12):
        ws2.cell(r, c).number_format = "#,##0.0000"
    for c in (13, 15):
        ws2.cell(r, c).number_format = "0.00%"

# ───────── Sheet 3: Monthly Returns ─────────
ws3 = wb.create_sheet("Monthly Returns")
hdr_r_top = [
    "Date",
    "Stocks", "Bonds", "Gold", "Managed Futures", "Futures Yield", "Merger Arb", "T-Bills",
    "Stock/Bond Return", "Stacked Return", "Difference (Stacked - S/B)",
    # intermediate components for full traceability
    "Trend Financed", "FY Net", "Gold Net", "Merger Arb Net",
]
hdr_r_sub = [
    "From underlying levels in 'Index Levels'",
    "B/B_prev - 1", "C/C_prev - 1", "D/D_prev - 1", "E/E_prev - 1", "F/F_prev - 1", "G/G_prev - 1", "H/H_prev - 1",
    "sw*r(Stk) + bw*r(Bnd)",
    "r(S/B) + ss * Sum(w_i * net_i)",
    "Column J - Column I",
    "r(Trend) - trendFee/12 - (r(Tbill) + trendFin/12)",
    "r(FY) - fyFee/12 - (r(Tbill) + fyFin/12)",
    "r(Gold) - goldFee/12 - (r(Tbill) + goldFin/12)",
    "r(MA) - marbFee/12 - (r(Tbill) + marbFin/12)",
]
for c, h in enumerate(hdr_r_top, 1):
    cell = ws3.cell(1, c, h); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = header_fill
for c, h in enumerate(hdr_r_sub, 1):
    ws3.cell(2, c, h).font = italic

for i in range(1, n):
    r = 2 + i               # return row r corresponds to month i (i=1 -> row 3)
    src_curr = 3 + i        # Index Levels row for current month
    src_prev = 2 + i        # Index Levels row for prior month
    ws3.cell(r, 1, DATA[i]["date"])
    for c_off, col in enumerate(["B", "C", "D", "E", "F", "G", "H"]):
        ws3.cell(r, 2 + c_off,
                 f"='Index Levels'!{col}{src_curr}/'Index Levels'!{col}{src_prev}-1")
    # Stock/Bond return
    ws3.cell(r, 9, f"=(sw/100)*B{r}+(bw/100)*C{r}")
    # Net / financed per overlay
    ws3.cell(r, 12, f"=E{r}-trendFee/10000/12-(H{r}+trendFinancing/10000/12)")
    ws3.cell(r, 13, f"=F{r}-fyFee/10000/12-(H{r}+fyFinancing/10000/12)")
    ws3.cell(r, 14, f"=D{r}-goldFee/10000/12-(H{r}+goldFinancing/10000/12)")
    ws3.cell(r, 15, f"=G{r}-marbFee/10000/12-(H{r}+marbFinancing/10000/12)")
    # Stacked return
    ws3.cell(r, 10, f"=I{r}+(ss/100)*((tw/100)*L{r}+(fw/100)*M{r}+(gw/100)*N{r}+(mw/100)*O{r})")
    # Difference
    ws3.cell(r, 11, f"=J{r}-I{r}")

widths3 = [12, 14, 14, 14, 20, 18, 20, 14, 20, 20, 26, 16, 14, 14, 18]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "B3"
returns_last_row = 2 + n_returns
for r in range(3, returns_last_row + 1):
    for c in range(2, 16):
        ws3.cell(r, c).number_format = "0.0000%"

# ───────── Sheet 4: Statistics ─────────
ws4 = wb.create_sheet("Statistics")
ws4["A1"] = "Summary Statistics (formulas mirror the widget's compute logic)"
ws4["A1"].font = Font(bold=True, size=12)

ws4.cell(3, 1, "Metric").font = bold
ws4.cell(3, 2, "Stock/Bond").font = Font(bold=True, color="FFFFFF"); ws4.cell(3, 2).fill = header_fill
ws4.cell(3, 3, "Stacked").font = Font(bold=True, color="FFFFFF"); ws4.cell(3, 3).fill = header_fill
ws4.cell(3, 4, "Difference (Stacked - S/B)").font = Font(bold=True, color="FFFFFF"); ws4.cell(3, 4).fill = header_fill

I_first = "'Index Levels'!I3"
I_last  = f"'Index Levels'!I{last_row_levels}"
J_first = "'Index Levels'!J3"
J_last  = f"'Index Levels'!J{last_row_levels}"
I_rets  = f"'Monthly Returns'!I3:I{returns_last_row}"
J_rets  = f"'Monthly Returns'!J3:J{returns_last_row}"
K_rets  = f"'Monthly Returns'!K3:K{returns_last_row}"
H_rets  = f"'Monthly Returns'!H3:H{returns_last_row}"
DD_I    = f"'Index Levels'!M3:M{last_row_levels}"
DD_J    = f"'Index Levels'!O3:O{last_row_levels}"

# Annualized Return
ws4.cell(4, 1, "Annualized Return")
ws4.cell(4, 2, f"=({I_last}/{I_first})^(12/{n_returns})-1")
ws4.cell(4, 3, f"=({J_last}/{J_first})^(12/{n_returns})-1")
ws4.cell(4, 4, "=C4-B4")

# Annualized Volatility
ws4.cell(5, 1, "Annualized Volatility")
ws4.cell(5, 2, f"=STDEV.S({I_rets})*SQRT(12)")
ws4.cell(5, 3, f"=STDEV.S({J_rets})*SQRT(12)")
ws4.cell(5, 4, "=C5-B5")

# Max Drawdown
ws4.cell(6, 1, "Maximum Drawdown")
ws4.cell(6, 2, f"=MAX({DD_I})")
ws4.cell(6, 3, f"=MAX({DD_J})")
ws4.cell(6, 4, "=C6-B6")

# Sharpe Ratio
ws4.cell(7, 1, "Sharpe Ratio")
ws4.cell(7, 2, f"=(B4-AVERAGE({H_rets})*12)/B5")
ws4.cell(7, 3, f"=(C4-AVERAGE({H_rets})*12)/C5")
ws4.cell(7, 4, "=C7-B7")

# Tracking Error
ws4.cell(8, 1, "Annualized Tracking Error (vs Stock/Bond)")
ws4.cell(8, 2, 0)
ws4.cell(8, 3, f"=STDEV.S({K_rets})*SQRT(12)")
ws4.cell(8, 4, "=C8-B8")

for rr in (4, 5, 6, 8):
    for c in (2, 3, 4):
        ws4.cell(rr, c).number_format = "0.00%"
for c in (2, 3, 4):
    ws4.cell(7, c).number_format = "0.000"

ws4.column_dimensions["A"].width = 42
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 18
ws4.column_dimensions["D"].width = 28

# Methodology footer
ws4.cell(10, 1, "Methodology (mirrors Simple Visualizer compute() and disclosures)").font = bold
methodology = [
    "  Annualized Return  = (End Level / Start Level)^(12 / n_returns) - 1",
    "  Annualized Volatility = STDEV.S(monthly returns) * SQRT(12)",
    "  Maximum Drawdown   = MAX over t of (Peak(0..t) - Level(t)) / Peak(0..t)",
    "  Sharpe Ratio       = (Ann Return - AVERAGE(T-Bill monthly returns)*12) / Ann Vol",
    "  Annualized Tracking Error = STDEV.S(Stacked - Stock/Bond) * SQRT(12)",
    "",
    "Stacked return formula (per Simple Visualizer compute()):",
    "  R_Fixed   = sw * r(Stocks) + bw * r(Bonds)",
    "  R_Trend_F = r(Trend) - trendFee/12 - (r(T-Bills) + trendFinancing/12)",
    "  R_FY_N    = r(FY)    - fyFee/12    - (r(T-Bills) + fyFinancing/12)",
    "  R_Gold_N  = r(Gold)  - goldFee/12  - (r(T-Bills) + goldFinancing/12)",
    "  R_MArb_N  = r(MArb)  - marbFee/12  - (r(T-Bills) + marbFinancing/12)",
    "  R_Stacked = R_Fixed + ss * (tw*R_Trend_F + fw*R_FY_N + gw*R_Gold_N + mw*R_MArb_N)",
    "  All weights divided by 100 in Excel formulas (Inputs are stored as percentages).",
    "  Fees and financing inputs are basis points; converted via /10000/12 to a monthly decimal.",
]
for i, line in enumerate(methodology, 11):
    ws4.cell(i, 1, line)

# ───────── Sheet 5: Notes / Disclosures ─────────
ws5 = wb.create_sheet("Notes")
notes = [
    "Source: live Return Stacked Portfolio Visualizer (returnstacked.com)",
    "  Script: /wp-content/uploads/visualizer-widget/visualizer-widget.js",
    "Data, formulas, and disclosures reproduced verbatim for compliance documentation.",
    "",
    "Underlying indices (Index Levels columns B-H):",
    "  Stocks               S&P 500 Total Return Index (SPX)",
    "  Bonds                Bloomberg US Aggregate Bond Index (LBUSTRUU)",
    "  Gold                 Spot Gold quoted in US Dollars (XAU Currency)",
    "  Managed Futures      PivotalPath Managed Futures Index (MFT) - reported net of fees",
    "  Futures Yield        Futures Yield (Carry) Index",
    "  Merger Arbitrage     PivotalPath Event Driven: Merger Arbitrage Index (EVDMER) - net of fees",
    "  T-Bills              Bloomberg Short Treasury US Total Return Index (LD12TRUUU)",
    "",
    "Disclosures (from Simple Visualizer):",
    "  - Bloomberg Short Treasury US Total Return Index tracks the market for Treasury bills",
    "    issued by the US government with time to maturity between 1 and 3 months.",
    "  - Bloomberg US Aggregate Bond Index covers the broad U.S. investment grade, USD-denominated,",
    "    fixed-rate taxable bond market.",
    "  - S&P 500 Index is a market-capitalization-weighted index of 500 leading publicly traded",
    "    companies in the U.S.",
    "  - PivotalPath indices (MFT, EVDMER) are reported net of fees.",
    "",
    "Workbook structure:",
    "  Inputs           Editable weights, fees, financing (cells highlighted yellow).",
    "                   Defaults match the simple visualizer defaults (40 bp Gold fee,",
    "                   50 bp financing on every overlay) and pre-fill the user's stated",
    "                   60/40 + 30% Stack scheme (33.34% MF / 33.33% MA / 33.33% Gold).",
    "  Index Levels     Raw underlying price levels (cols B-H) and derived portfolio",
    "                   level series (cols I-K).  Helper columns L-O drive max drawdown.",
    "  Monthly Returns  Per-month returns of underlying series (formula = level ratio - 1)",
    "                   and derived portfolio returns including the intermediate net/",
    "                   financed components per overlay (cols L-O).",
    "  Statistics       Five-metric table with formula cells: Annualized Return, Vol,",
    "                   Max Drawdown, Sharpe Ratio, Annualized Tracking Error.",
    "",
    f"Data span: {DATA[0]['date']} to {DATA[-1]['date']} ({n} monthly observations, {n_returns} returns).",
    "Base date for derived growth-of-100 series: " + DATA[0]["date"] + ".",
]
for i, line in enumerate(notes, 1):
    ws5.cell(i, 1, line)
ws5.column_dimensions["A"].width = 110

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  Index Levels rows: 3..{last_row_levels} ({n} months)")
print(f"  Monthly Returns rows: 3..{returns_last_row} ({n_returns} returns)")
print(f"  Inputs: 15 editable cells with defined names (sw, bw, ss, tw, fw, gw, mw, fees, financings)")
