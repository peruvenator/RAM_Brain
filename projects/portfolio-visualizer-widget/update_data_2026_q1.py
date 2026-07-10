"""
Splice 2026 Q1 (Jan / Feb / Mar) into data/indices_compact.json.

Sources:
- Index Data to 2026-03-31.xlsx ........ 54 named series, full history
- Futures Yield Q1 2026.csv ............ monthly returns
- Treasury Ladder to 2025-03-31.csv .... monthly returns
- Alphabeta merger arbitrage to 2026-04-30.csv ... monthly returns

Rules:
- Append 3 new month-end dates to `dates`: 2026-01-31, 2026-02-28, 2026-03-31.
- For 8 "materially different" series, REPLACE the full history with the new file's series.
- For all other series in the new file (including 2 renames), APPEND only the new 3 months.
- For 3 series not in the new file, chain-link the CSV returns onto the existing last level.
"""
import json
import csv
import openpyxl
from datetime import datetime
from copy import deepcopy

EXCEL_PATH = "Index Data to 2026-03-31.xlsx"
INDICES_PATH = "data/indices_compact.json"
FUTURES_YIELD_PATH = "data/futures_yield.json"
NEW_DATES = ["2026-01-31", "2026-02-28", "2026-03-31"]

# Series to wholesale-replace (history disagrees with current widget data).
REPLACE_SERIES = {
    "CRDREL",
    "EQDGLS",
    "EVD",
    "EurekaHedge CBOE Long Vol",
    "EurekaHedge CBOE Tail Risk",
    "EurekaHedge Equity Market Neutral Hedge Fund Index",
    "HFCEW",
    "GBM",
}

# Widget keeps these names; the new file uses the Bloomberg label.
RENAME_MAP = {
    "U.S. Treasuries": "Bloomberg US Treasury Index",
    "U.S. Corporate Fixed Income": "Bloomberg US Corporate Index",
}

CSV_SERIES = [
    ("Futures Yield (Carry)", "Futures Yield Q1 2026.csv"),
    ("U.S. Treasury Ladder", "Treasury Ladder to 2025-03-31.csv"),
    ("AlphaBeta Merger Arbitrage Index", "Alphabeta merger arbitrage to 2026-04-30.csv"),
]


def load_new_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Sheet1"]
    headers = {}
    for cell in ws[1]:
        if cell.value and cell.column >= 2:
            headers[cell.column] = cell.value
    series = {name: [] for name in headers.values()}
    dates = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        d = row[0].value
        if not isinstance(d, datetime):
            continue
        dates.append(d.strftime("%Y-%m-%d"))
        for col, name in headers.items():
            v = row[col - 1].value
            if v == "" or v is None:
                series[name].append(None)
            else:
                try:
                    series[name].append(float(v))
                except (TypeError, ValueError):
                    series[name].append(None)
    return dates, series


def load_csv_returns(path):
    """Return list of floats (decimal, not percent) for Jan/Feb/Mar 2026."""
    rows = []
    with open(path, newline="") as f:
        reader = csv.reader(f)
        next(reader)  # header
        for r in reader:
            if not r or not r[0].strip() or not r[1].strip():
                continue
            rows.append(r[1].strip())
    if len(rows) < 3:
        raise ValueError(f"{path}: expected 3 monthly return rows, got {len(rows)}")
    out = []
    for v in rows[:3]:
        v = v.replace("%", "").strip()
        out.append(float(v) / 100.0)
    return out


def chain_link(last_level, returns):
    out = []
    cur = last_level
    for r in returns:
        cur = cur * (1.0 + r)
        out.append(cur)
    return out


def dense(series_obj, total_len):
    arr = [None] * total_len
    for i, v in enumerate(series_obj["values"]):
        arr[series_obj["start"] + i] = v
    return arr


def compact(values):
    start = 0
    while start < len(values) and values[start] is None:
        start += 1
    if start == len(values):
        return {"start": 0, "values": []}
    end = len(values) - 1
    while end > start and values[end] is None:
        end -= 1
    return {"start": start, "values": values[start:end + 1]}


def main():
    print("Loading existing indices_compact.json...")
    with open(INDICES_PATH) as f:
        existing = json.load(f)

    existing_dates = existing["dates"]
    existing_len = len(existing_dates)
    print(f"  {existing_len} dates, {len(existing['series'])} series")
    assert existing_dates[-1] == "2025-12-31", f"unexpected last date {existing_dates[-1]}"

    print(f"Loading {EXCEL_PATH}...")
    new_dates, new_series = load_new_excel()
    new_len = len(new_dates)
    print(f"  {new_len} dates, {len(new_series)} series")
    assert new_dates[-3:] == NEW_DATES, f"new file tail doesn't match: {new_dates[-3:]}"
    assert new_len == existing_len + 3, f"row count mismatch: existing {existing_len}, new {new_len}"

    # New dates array.
    out_dates = existing_dates + NEW_DATES
    out_len = len(out_dates)

    out_series = {}

    # 1. Replace 8 series wholesale from new file.
    replaced = 0
    for name in REPLACE_SERIES:
        if name not in existing["series"]:
            raise KeyError(f"REPLACE target {name!r} not in existing")
        if name not in new_series:
            raise KeyError(f"REPLACE source {name!r} not in new file")
        out_series[name] = compact(new_series[name])
        replaced += 1
    print(f"Replaced: {replaced}")

    # 2. Renames: append new 3 months from differently-named new-file columns.
    renamed = 0
    for old_name, new_name in RENAME_MAP.items():
        if old_name not in existing["series"]:
            raise KeyError(f"RENAME target {old_name!r} not in existing")
        if new_name not in new_series:
            raise KeyError(f"RENAME source {new_name!r} not in new file")
        # Verify last overlap value matches before appending.
        d = dense(existing["series"][old_name], existing_len)
        e_last, n_last = d[-1], new_series[new_name][existing_len - 1]
        if e_last is not None and n_last is not None and abs(e_last - n_last) / abs(e_last) > 1e-3:
            raise ValueError(f"RENAME {old_name}<->{new_name} mismatch at last overlap: {e_last} vs {n_last}")
        appended_values = d + new_series[new_name][-3:]
        out_series[old_name] = compact(appended_values)
        renamed += 1
    print(f"Renamed (append from differently-named column): {renamed}")

    # 3. CSV-sourced series: chain-link returns onto last level.
    csv_done = 0
    for name, csv_path in CSV_SERIES:
        if name not in existing["series"]:
            raise KeyError(f"CSV target {name!r} not in existing")
        rets = load_csv_returns(csv_path)
        d = dense(existing["series"][name], existing_len)
        last_level = d[-1]
        if last_level is None:
            raise ValueError(f"{name}: existing last level is None")
        new_levels = chain_link(last_level, rets)
        appended_values = d + new_levels
        out_series[name] = compact(appended_values)
        print(f"  {name}: last={last_level:.4f} -> {new_levels[0]:.4f}, {new_levels[1]:.4f}, {new_levels[2]:.4f}")
        csv_done += 1
    print(f"CSV-sourced (chain-linked): {csv_done}")

    # 4. All other series in new file: chain-link 3 new monthly returns onto
    #    existing last level. This preserves continuity even when the new file
    #    rebases a series (EVDMER and MFT in the new file are scaled ~10x off).
    handled = set(out_series.keys())
    already_seen_new_names = set(RENAME_MAP.values()) | REPLACE_SERIES
    appended = 0
    for name in new_series:
        if name in already_seen_new_names:
            continue
        if name not in existing["series"]:
            print(f"  NOTE: new-file series {name!r} not in existing widget; skipping")
            continue
        d = dense(existing["series"][name], existing_len)
        n_vals = new_series[name]
        # Compute 3 monthly returns from new file across positions (existing_len-1, existing_len-1+3)
        base_idx = existing_len - 1  # position of last shared date in new file
        rets = []
        for i in range(1, 4):
            prev = n_vals[base_idx + i - 1]
            cur = n_vals[base_idx + i]
            if prev is None or cur is None or prev == 0:
                rets.append(None)
            else:
                rets.append(cur / prev - 1.0)
        e_last = d[-1]
        if e_last is None or any(r is None for r in rets):
            print(f"  SKIP append {name}: insufficient data (existing_last={e_last}, rets={rets})")
            out_series[name] = compact(d + [None, None, None])
            continue
        new_levels = chain_link(e_last, rets)
        out_series[name] = compact(d + new_levels)
        # Flag noticeable scale rebases (informational only).
        n_last = n_vals[base_idx]
        if n_last and abs(e_last - n_last) / abs(e_last) > 1e-3:
            print(f"  INFO: {name} rebased in source (existing={e_last:.4f}, new={n_last:.4f}); chained returns instead")
        appended += 1
    print(f"Chain-linked (3 months from new-file returns): {appended}")

    # 5. Any existing series not yet handled (should be none): pad with 3 Nones.
    leftover = [s for s in existing["series"] if s not in out_series]
    for s in leftover:
        print(f"  UNHANDLED existing series (padding nulls): {s}")
        d = dense(existing["series"][s], existing_len)
        out_series[s] = compact(d + [None, None, None])

    out = {"dates": out_dates, "series": out_series}

    # Sanity: every series values+start must fit within out_len.
    for name, s in out_series.items():
        end = s["start"] + len(s["values"])
        if end > out_len:
            raise ValueError(f"series {name} overruns: end={end} > {out_len}")

    print(f"\nWriting {INDICES_PATH} ...")
    with open(INDICES_PATH, "w") as f:
        json.dump(out, f)
    print(f"  dates={out_len}, series={len(out_series)}")

    # 6. Also extend data/futures_yield.json so the standalone artifact stays in sync.
    try:
        with open(FUTURES_YIELD_PATH) as f:
            fy = json.load(f)
        fy_rets = load_csv_returns("Futures Yield Q1 2026.csv")
        last = fy["values"][-1]
        new_vals = chain_link(last, fy_rets)
        fy["dates"].extend(NEW_DATES)
        fy["values"].extend(new_vals)
        with open(FUTURES_YIELD_PATH, "w") as f:
            json.dump(fy, f)
        print(f"  updated futures_yield.json: now {len(fy['dates'])} rows")
    except FileNotFoundError:
        print("  futures_yield.json not present, skipping")

    print("\nDone.")


if __name__ == "__main__":
    main()
